#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ab_power_meter_monitor.py
=========================
Allen-Bradley (Rockwell) Site Power Meter — Web Table Poller & Data Logger
NRAO / GBO Site Infrastructure Monitoring Tool

Polls Allen-Bradley power meters over HTTP, parses the 11 HTML data tables
(pages 0-10), stores data in named Python dicts, and writes results to any
combination of:
  • Live PyQt/PySide6 GUI with embedded matplotlib preview plots
  • NRAO-compliant FITS files
  • Per-table CSV files (columnar: timestamp=row, parameter=column)
  • Excel XLSX workbook (one sheet per table, line charts over time)
  • Appending columnar Markdown log files (.md) — GFM pipe tables, one file
    per device per table (timestamp=row, parameter=column)
  • Veusz HDF5 (.vszh5) project files — written at loop end with all samples

IP address base: 10.16.130.{last_octet}
Device range   : last octet 50 – 53  (configurable below)

Author : W. Wallace — NRAO / Green Bank Observatory
Date   : 2026-05-11
Python : 3.8+
Deps   : PySide6, matplotlib, requests, beautifulsoup4, lxml,
         astropy, openpyxl, veusz  
         (pip, conda, mamba, or uv install each)

Usage
-----
Headless / scripted:
    python ab_power_meter_monitor.py

GUI mode (set ENABLE_GUI = 1 below or check the checkbox at launch):
    python ab_power_meter_monitor.py   # then toggle via GUI

All output-enable switches can be overridden at runtime via the GUI.

Notes on the HTML endpoint
---------------------------
Each meter exposes 11 table pages at:
    http://<IP>/<page_index>
where page_index is 0 … 10.  The <body> tag contains a single <table>
with rows: # | Parameter Name | Value.

Table map (page index → human name):
    0  Device Configuration Table
    1  Communications Configuration Table
    2  Voltage / Current Table
    3  Real Time Power Table
    4  Cumulative Power Table
    5  Demand Data Table
    6  Diagnostic Table
    7  Voltage / Current Snapshot Log Table
    8  Power Snapshot Log Table
    9  Min_Max Log Table
   10  Diagnostic Table (extended)
"""

# ===========================================================================
#  STANDARD-LIBRARY IMPORTS
# ===========================================================================
import os
import sys
import csv
import json
import time
import logging
import datetime
import traceback
from typing import Any, Dict, List, Optional, Tuple
import threading
import signal
import concurrent.futures
try:
    import psutil          # pip install psutil — for memory monitoring
except ImportError:
    # type: ignore[assignment]  graceful degradation: memory-based
    psutil = None
    # flush threshold will not trigger, sentinel 9999 MB is returned

# ===========================================================================
#  ██████╗  ██████╗ ██╗    ██╗███████╗██████╗     ███████╗██╗    ██╗██╗████████╗ ██████╗██╗  ██╗███████╗███████╗
#  ██╔══██╗██╔═══██╗██║    ██║██╔════╝██╔══██╗    ██╔════╝██║    ██║██║╚══██╔══╝██╔════╝██║  ██║██╔════╝██╔════╝
#  ██████╔╝██║   ██║██║ █╗ ██║█████╗  ██████╔╝    ███████╗██║ █╗ ██║██║   ██║   ██║     ███████║█████╗  ███████╗
#  ██╔═══╝ ██║   ██║██║███╗██║██╔══╝  ██╔══██╗    ╚════██║██║███╗██║██║   ██║   ██║     ██╔══██║██╔══╝  ╚════██║
#  ██║     ╚██████╔╝╚███╔███╔╝███████╗██║  ██║    ███████║╚███╔███╔╝██║   ██║   ╚██████╗██║  ██║███████╗███████║
#  ╚═╝      ╚═════╝  ╚══╝╚══╝ ╚══════╝╚═╝  ╚═╝    ╚══════╝ ╚══╝╚══╝ ╚═╝   ╚═╝    ╚═════╝╚═╝  ╚═╝╚══════╝╚══════╝
#
#  ALL RUNTIME BEHAVIOUR IS CONTROLLED BY THE VARIABLES IN THIS SECTION.
#  Set 0 = False / disabled,  1 = True / enabled.
# ===========================================================================

# ---------------------------------------------------------------------------
# Feature / output enable switches  (0 = off, 1 = on)
# ---------------------------------------------------------------------------
ENABLE_GUI = 0   # Show PyQt/PySide6 main window
ENABLE_FITS = 1   # Write NRAO-compliant FITS files
ENABLE_CSV = 1   # Write per-table CSV files
ENABLE_XLSX = 1   # Write Excel workbook with charts
# Append timestamped entries to Markdown log files (.md)
ENABLE_LOG_APPEND = 1
ENABLE_VEUSZ = 1   # Write Veusz HDF5 project file(s) (.vszh5)

# ---------------------------------------------------------------------------
# Headless loop control
# ---------------------------------------------------------------------------
# Number of poll cycles to run in headless mode.
# 0 = run indefinitely until stopped (Ctrl-C or stop signal file).
# N = run exactly N cycles then exit cleanly.
HEADLESS_LOOP_COUNT = 100   # 0 = infinite loop; N = run N cycles then stop

# ---------------------------------------------------------------------------
# Memory / flush thresholds (adaptive write scheduling)
# ---------------------------------------------------------------------------
# When the in-memory time-series store grows beyond MEM_FLUSH_THRESHOLD_MB
# OR free system RAM drops below MEM_FREE_MIN_MB, an intermediate flush of
# CSV / XLSX / log files is triggered mid-loop (parallel, non-blocking) so
# memory is reclaimed without dropping sample points.
MEM_FLUSH_THRESHOLD_MB = 256   # flush when store occupies more than N MB
MEM_FREE_MIN_MB = 512   # flush when system free RAM falls below N MB

# ---------------------------------------------------------------------------
# IP address configuration
# ---------------------------------------------------------------------------
IP_BASE = "10.16.130"  # First three octets (do NOT include trailing dot)
IP_LAST_OCTET_START = 50           # Start of last-octet range (inclusive)
IP_LAST_OCTET_END = 51           # End   of last-octet range (inclusive)

# ---------------------------------------------------------------------------
# Polling / timing
# ---------------------------------------------------------------------------
SAMPLE_PERIOD_SEC = 15    # Seconds between successive polls of all devices
HTTP_TIMEOUT_SEC = 5     # Per-request HTTP timeout

# ---------------------------------------------------------------------------
# Output paths
# ---------------------------------------------------------------------------
# PRIMARY OUTPUT ROOT — change OUTPUT_BASE_DIR to redirect ALL output
# (logs, FITS, CSV, XLSX, Veusz) to a different location without touching
# any of the sub-directory constants below.
#
# Set to an absolute path to store output anywhere on the filesystem, e.g.:
#   OUTPUT_BASE_DIR = "/mnt/data/ab_meter_output"
#   OUTPUT_BASE_DIR = r"D:\GBO\PowerMeter\output"
#
# The default resolves to a folder named "ab_meter_output" sitting next
# to this script file, which keeps everything self-contained.
# ---------------------------------------------------------------------------
OUTPUT_BASE_DIR = os.path.join(os.path.dirname(
    os.path.abspath(__file__)), "ab_meter_output")

# Sub-directories — all derived from OUTPUT_BASE_DIR so a single change
# above propagates everywhere automatically.  Override individually only
# if you need outputs split across different locations.
# kept for back-compat references
OUTPUT_DIR = OUTPUT_BASE_DIR
LOG_DIR = os.path.join(OUTPUT_BASE_DIR, "logs")   # text log files
FITS_DIR = os.path.join(OUTPUT_BASE_DIR, "fits")   # NRAO FITS files
CSV_DIR = os.path.join(OUTPUT_BASE_DIR, "csv")    # per-table CSV files
XLSX_DIR = os.path.join(OUTPUT_BASE_DIR, "xlsx")   # Excel workbooks
VEUSZ_DIR = os.path.join(OUTPUT_BASE_DIR, "veusz")  # Veusz HDF5 projects

# Path to the stop-signal file.  Touch this file (or run ab_stop.py) to
# request a clean shutdown of the headless loop.  Deleted automatically
# on startup and on clean exit.
STOP_SIGNAL_FILE = os.path.join(OUTPUT_BASE_DIR, "STOP_COLLECTION")

# ---------------------------------------------------------------------------
# Table page-index → canonical name mapping
# ---------------------------------------------------------------------------
TABLE_NAMES: Dict[int, str] = {
    0:  "Device_Configuration_Table",
    1:  "Communications_Configuration_Table",
    2:  "Voltage_Current_Table",
    3:  "Real_Time_Power_Table",
    4:  "Cumulative_Power_Table",
    5:  "Demand_Data_Table",
    6:  "Diagnostic_Table",
    7:  "Voltage_Current_Snapshot_Log_Table",
    8:  "Power_Snapshot_Log_Table",
    9:  "MinMax_Log_Table",
    10: "Diagnostic_Table_Extended",
}

# ---------------------------------------------------------------------------
# Unit inference map: substring → unit label
# Applied when building Veusz axis labels and FITS column units.
# Keys are LOWER-CASE substrings found in parameter names.
# ---------------------------------------------------------------------------
UNIT_MAP: List[Tuple[str, str]] = [
    ("current",          "A"),
    ("voltage",          "V"),
    ("frequency",        "Hz"),
    ("kw hour",          "kWh"),
    ("kvar hour",        "kVARh"),
    ("real power",       "W"),
    ("reactive power",   "VAR"),
    ("apparent power",   "VA"),
    ("true pf",          "%"),
    ("displacement pf",  "%"),
    ("distortion pf",    "%"),
    ("demand current",   "A"),
    ("demand power",     "W"),
    ("demand apparent",  "VA"),
    ("demand reactive",  "VAR"),
    ("elapsed time",     "s"),
    ("period",           "min"),
    ("interval",         "s"),
    ("pulse width",      "ms"),
]


# ===========================================================================
#  LOGGING SETUP
# ===========================================================================
def _setup_logging(log_dir: str, append: bool = True) -> logging.Logger:
    """
    Initialise the module-level logger.

    Parameters
    ----------
    log_dir : str
        Directory where the rotating log file will be written.
    append : bool
        If True, append to existing log file; otherwise overwrite.

    Returns
    -------
    logging.Logger
        Configured logger instance.
    """
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "ab_monitor.log")
    file_mode = "a" if append else "w"

    fmt = logging.Formatter(
        fmt="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S",
    )

    logger = logging.getLogger("ABMonitor")
    logger.setLevel(logging.DEBUG)

    # File handler
    fh = logging.FileHandler(log_path, mode=file_mode, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    # Console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    if not logger.handlers:
        logger.addHandler(fh)
        logger.addHandler(ch)

    logger.info("Logger initialised — output: %s", log_path)
    return logger


# Global logger (module scope; re-initialised when log_dir changes)
logger: logging.Logger = _setup_logging(
    LOG_DIR, append=bool(ENABLE_LOG_APPEND))


# ===========================================================================
#  FITS ASCII SANITISER
# ===========================================================================
def _fits_ascii(value: str) -> str:
    """
    Sanitise a string so it contains only printable 7-bit ASCII characters
    safe for use in a FITS header keyword value, COMMENT, or HISTORY field.

    FITS standard (NOST 100-2.0, sect. 4.4.2) restricts header character
    values to ASCII bytes 0x20 – 0x7E.  Any character outside that range is
    replaced with a plain hyphen '-' so the keyword is never rejected by
    astropy or CFITSIO.

    Common culprits:
      - Unicode em-dash U+2014 (—)  → ' - '
      - Unicode en-dash U+2013 (–)  → ' - '
      - Degree sign, mu, etc.        → '-'

    Parameters
    ----------
    value : str
        Raw header value string that may contain non-ASCII characters.

    Returns
    -------
    str
        Pure printable ASCII string, max 68 characters (FITS card limit
        for a string value after the keyword and value indicator).
    """
    # Replace the most common typographic substitutes first for legibility
    value = value.replace("\u2014", " - ")   # em-dash
    value = value.replace("\u2013", " - ")   # en-dash
    value = value.replace("\u00b0", "deg")   # degree sign
    value = value.replace("\u03bc", "u")     # Greek mu (micro)
    value = value.replace("\u03a9", "Ohm")   # Greek capital omega
    # Replace any remaining non-printable or non-ASCII byte with '-'
    sanitised = "".join(
        c if (0x20 <= ord(c) <= 0x7E) else "-"
        for c in value
    )
    return sanitised[:68]


# ===========================================================================
#  UNIT INFERENCE HELPER
# ===========================================================================
def infer_unit(param_name: str) -> str:
    """
    Infer a physical unit string from a parameter name using UNIT_MAP.

    Parameters
    ----------
    param_name : str
        The human-readable parameter name string from the meter HTML table.

    Returns
    -------
    str
        Unit label string, e.g. 'A', 'V', 'W', or '' if unknown.
    """
    lower = param_name.lower()
    for key, unit in UNIT_MAP:
        if key in lower:
            return unit
    return ""


# ===========================================================================
#  HTML FETCH & PARSE
# ===========================================================================
def fetch_table_html(ip: str, page: int, timeout: int = HTTP_TIMEOUT_SEC) -> Optional[str]:
    """
    Fetch raw HTML for a single meter table page via HTTP GET.

    Parameters
    ----------
    ip : str
        Full IP address string, e.g. '10.16.130.50'.
    page : int
        Table page index (0–10).
    timeout : int
        HTTP request timeout in seconds.

    Returns
    -------
    Optional[str]
        HTML text body on success, None on any error.
    """
    import requests  # local import to keep headless-mode dependency optional

    url = f"http://{ip}/{page}"
    try:
        resp = requests.get(url, timeout=timeout)
        resp.raise_for_status()
        logger.debug("Fetched %s — %d bytes", url, len(resp.text))
        return resp.text
    except Exception as exc:  # broad catch intentional — network errors vary widely
        logger.warning("Failed to fetch %s: %s", url, exc)
        return None


def parse_html_table(html: str, table_name: str, ip: str, page: int) -> Dict[str, Any]:
    """
    Parse an Allen-Bradley power meter HTML table body into a Python dict.

    The HTML format is:
        <tr><td>#</td><td>Parameter Name</td><td>Value</td></tr>

    Parameters
    ----------
    html : str
        Raw HTML string from the meter.
    table_name : str
        Canonical table name used as the dict key prefix.
    ip : str
        Source IP address (stored as metadata).
    page : int
        Source page index (stored as metadata).

    Returns
    -------
    Dict[str, Any]
        Dictionary containing:
        - '_meta'  : dict  — source info, timestamp, table name
        - '#N_name': str   — parameter name (key = index string)
        - '#N_value': Any  — parsed numeric or string value
        - '#N_unit' : str  — inferred SI unit or ''
    """
    from bs4 import BeautifulSoup

    result: Dict[str, Any] = {
        "_meta": {
            "table_name":  table_name,
            "source_ip":   ip,
            "page_index":  page,
            "fetch_utc":   datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        }
    }

    if html is None:
        result["_meta"]["error"] = "No HTML received"
        return result

    try:
        soup = BeautifulSoup(html, "lxml")
        rows = soup.find_all("tr")
        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 3:
                continue  # skip header row (uses <font> not <td> data)
            idx_text = cells[0].get_text(strip=True)
            param_name = cells[1].get_text(strip=True)
            raw_value = cells[2].get_text(strip=True)

            # Skip header-looking rows
            if param_name in ("Parameter Name", "#") or idx_text == "#":
                continue

            # Attempt numeric conversion
            try:
                value: Any = float(raw_value)
                if value == int(value) and "." not in raw_value:
                    value = int(value)
            except (ValueError, TypeError):
                value = raw_value  # keep as string (dates, '####', etc.)

            key = f"#{idx_text}"
            result[key] = param_name
            result[f"{key}_value"] = value
            result[f"{key}_unit"] = infer_unit(param_name)

        logger.debug("Parsed table '%s' — %d rows",
                     table_name, (len(result) - 1) // 3)

    except Exception as exc:
        logger.error("Parse error for table '%s': %s\n%s",
                     table_name, exc, traceback.format_exc())
        result["_meta"]["error"] = str(exc)

    return result


def poll_all_devices(
    ip_base: str,
    octet_start: int,
    octet_end: int,
    table_names: Dict[int, str],
) -> Dict[str, Dict[str, Any]]:
    """
    Poll all tables from all devices in the IP range.

    For each device (last-octet from octet_start to octet_end inclusive)
    and each of the 11 table pages, fetch and parse the HTML.

    Parameters
    ----------
    ip_base : str
        First three IP octets, e.g. '10.16.130'.
    octet_start : int
        First value of last octet to poll.
    octet_end : int
        Last  value of last octet to poll (inclusive).
    table_names : Dict[int, str]
        Mapping of page index → canonical table name.

    Returns
    -------
    Dict[str, Dict[str, Any]]
        Top-level key = "<IP>_<table_name>", value = parsed data dict.
        There will be at least 11 dicts per device.
    """
    all_data: Dict[str, Dict[str, Any]] = {}

    for last_octet in range(octet_start, octet_end + 1):
        ip = f"{ip_base}.{last_octet}"
        logger.info("Polling device %s …", ip)

        for page_idx, tname in table_names.items():
            dict_key = f"{ip}_{tname}"
            html = fetch_table_html(ip, page_idx)
            parsed = parse_html_table(html, tname, ip, page_idx)
            all_data[dict_key] = parsed
            logger.debug("Stored dict key: %s", dict_key)

    return all_data


# ===========================================================================
#  NAMED TABLE DICTS  (always populated; used by all output modules)
#
#  These 11 module-level dicts correspond to the 11 meter pages.
#  They are populated by update_named_dicts() after each poll.
#  Consumer code should reference these dicts directly.
# ===========================================================================

# --- Device 1 (last octet = 50, placeholder; populated at runtime) ---
Device_Configuration_Table:            Dict[str, Any] = {}
Communications_Configuration_Table:    Dict[str, Any] = {}
Voltage_Current_Table:                 Dict[str, Any] = {}
Real_Time_Power_Table:                 Dict[str, Any] = {}
Cumulative_Power_Table:                Dict[str, Any] = {}
Demand_Data_Table:                     Dict[str, Any] = {}
Diagnostic_Table:                      Dict[str, Any] = {}
Voltage_Current_Snapshot_Log_Table:    Dict[str, Any] = {}
Power_Snapshot_Log_Table:              Dict[str, Any] = {}
MinMax_Log_Table:                      Dict[str, Any] = {}
Diagnostic_Table_Extended:             Dict[str, Any] = {}

# Multi-device storage: keyed by IP then table name
ALL_DEVICE_DATA: Dict[str, Dict[str, Dict[str, Any]]] = {}

# ===========================================================================
#  TIME-SERIES ACCUMULATOR
#  Columnar store: TIME_SERIES_STORE[ip][table_name] = {
#      "timestamps_local": [str, ...],   # local-time strings, one per poll
#      "columns":          {param: [val, ...]},  # growing list per parameter
#      "units":            {param: unit_str},    # static, set on first poll
#  }
#  This is the primary source for CSV / XLSX / log / Veusz output.
#  ALL_DEVICE_DATA is still updated each poll (latest snapshot) for FITS and
#  GUI preview.  The accumulator is the authoritative multi-sample record.
# ===========================================================================
TIME_SERIES_STORE: Dict[str, Dict[str, Dict[str, Any]]] = {}
_TS_LOCK = threading.Lock()   # protects TIME_SERIES_STORE across threads


def accumulate_poll(all_device_data: Dict[str, Dict[str, Dict[str, Any]]]) -> None:
    """
    Append the current poll snapshot to the TIME_SERIES_STORE.

    Called immediately after update_named_dicts() on every poll cycle.
    Thread-safe via _TS_LOCK.

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: {ip: {table_name: parsed_dict}} (latest snapshot).
    """
    local_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with _TS_LOCK:
        for ip, tables in all_device_data.items():
            if ip not in TIME_SERIES_STORE:
                TIME_SERIES_STORE[ip] = {}

            for tname, tdict in tables.items():
                series = extract_numeric_series(tdict)
                if not series:
                    continue

                if tname not in TIME_SERIES_STORE[ip]:
                    # First poll: initialise columns and units
                    TIME_SERIES_STORE[ip][tname] = {
                        "timestamps_local": [],
                        "columns": {param: [] for param in series},
                        "units":   {param: unit for param, (_, unit) in series.items()},
                    }
                else:
                    # Subsequent polls: add any newly-appearing parameters
                    existing = TIME_SERIES_STORE[ip][tname]["columns"]
                    for param in series:
                        if param not in existing:
                            # Back-fill with None for prior rows
                            n_existing = len(
                                TIME_SERIES_STORE[ip][tname]["timestamps_local"])
                            existing[param] = [None] * n_existing
                            TIME_SERIES_STORE[ip][tname]["units"][param] = series[param][1]

                tstore = TIME_SERIES_STORE[ip][tname]
                tstore["timestamps_local"].append(local_ts)
                for param, (val, _) in series.items():
                    tstore["columns"][param].append(val)
                # Fill None for any column not present in this sample
                n_ts = len(tstore["timestamps_local"])
                for col_list in tstore["columns"].values():
                    if len(col_list) < n_ts:
                        col_list.append(None)

    logger.debug("Accumulator updated — %d devices, local ts: %s",
                 len(all_device_data), local_ts)


def ts_store_size_mb() -> float:
    """
    Estimate the current memory footprint of TIME_SERIES_STORE in megabytes.

    Uses sys.getsizeof recursively on all nested lists and dicts.  This is
    an approximation — actual RSS impact may be higher due to Python object
    overhead.

    Returns
    -------
    float
        Estimated size in megabytes.
    """
    import sys

    def _size(obj: Any, seen: Optional[set] = None) -> int:
        if seen is None:
            seen = set()
        obj_id = id(obj)
        if obj_id in seen:
            return 0
        seen.add(obj_id)
        size = sys.getsizeof(obj)
        if isinstance(obj, dict):
            size += sum(_size(k, seen) + _size(v, seen)
                        for k, v in obj.items())
        elif isinstance(obj, (list, tuple)):
            size += sum(_size(i, seen) for i in obj)
        return size

    with _TS_LOCK:
        return _size(TIME_SERIES_STORE) / (1024 * 1024)


def system_free_ram_mb() -> float:
    """
    Return the amount of free system RAM in megabytes.

    Falls back to a large sentinel value (9999) if psutil is unavailable
    so that the flush threshold is never falsely triggered.

    Returns
    -------
    float
        Free RAM in megabytes, or 9999 if psutil is not installed.
    """
    try:
        import psutil as _ps
        return _ps.virtual_memory().available / (1024 * 1024)
    except ImportError:
        return 9999.0


def should_flush(cfg: Dict[str, Any]) -> bool:
    """
    Decide whether an intermediate mid-loop file flush should occur.

    Returns True when either:
      • TIME_SERIES_STORE footprint exceeds MEM_FLUSH_THRESHOLD_MB, or
      • System free RAM is below MEM_FREE_MIN_MB.

    Parameters
    ----------
    cfg : Dict
        Runtime config dict (reads flush threshold keys).

    Returns
    -------
    bool
    """
    threshold = cfg.get("mem_flush_threshold_mb", MEM_FLUSH_THRESHOLD_MB)
    free_min = cfg.get("mem_free_min_mb",        MEM_FREE_MIN_MB)
    store_mb = ts_store_size_mb()
    free_mb = system_free_ram_mb()
    if store_mb > threshold:
        logger.info(
            "Flush triggered: store size %.1f MB > threshold %.1f MB", store_mb, threshold)
        return True
    if free_mb < free_min:
        logger.info(
            "Flush triggered: free RAM %.1f MB < minimum %.1f MB", free_mb, free_min)
        return True
    return False


def update_named_dicts(all_data: Dict[str, Dict[str, Any]]) -> None:
    """
    Update module-level named dicts from the raw all_data poll result.

    This function also rebuilds ALL_DEVICE_DATA which groups data by IP.

    Parameters
    ----------
    all_data : Dict[str, Dict[str, Any]]
        Output of poll_all_devices().
    """
    global Device_Configuration_Table, Communications_Configuration_Table
    global Voltage_Current_Table, Real_Time_Power_Table, Cumulative_Power_Table
    global Demand_Data_Table, Diagnostic_Table, Voltage_Current_Snapshot_Log_Table
    global Power_Snapshot_Log_Table, MinMax_Log_Table, Diagnostic_Table_Extended
    global ALL_DEVICE_DATA

    ALL_DEVICE_DATA.clear()

    for dict_key, data in all_data.items():
        meta = data.get("_meta", {})
        ip = meta.get("source_ip", "unknown")
        tname = meta.get("table_name", "unknown")

        if ip not in ALL_DEVICE_DATA:
            ALL_DEVICE_DATA[ip] = {}
        ALL_DEVICE_DATA[ip][tname] = data

    # Populate module-level dicts from the FIRST available device
    # (convenience reference; full multi-device access via ALL_DEVICE_DATA)
    first_ip = next(iter(ALL_DEVICE_DATA), None)
    if first_ip is None:
        return

    dev = ALL_DEVICE_DATA[first_ip]
    Device_Configuration_Table = dev.get(
        "Device_Configuration_Table",          {})
    Communications_Configuration_Table = dev.get(
        "Communications_Configuration_Table",  {})
    Voltage_Current_Table = dev.get("Voltage_Current_Table",               {})
    Real_Time_Power_Table = dev.get("Real_Time_Power_Table",               {})
    Cumulative_Power_Table = dev.get("Cumulative_Power_Table",              {})
    Demand_Data_Table = dev.get("Demand_Data_Table",                   {})
    Diagnostic_Table = dev.get("Diagnostic_Table",                    {})
    Voltage_Current_Snapshot_Log_Table = dev.get(
        "Voltage_Current_Snapshot_Log_Table",  {})
    Power_Snapshot_Log_Table = dev.get(
        "Power_Snapshot_Log_Table",            {})
    MinMax_Log_Table = dev.get("MinMax_Log_Table",                    {})
    Diagnostic_Table_Extended = dev.get(
        "Diagnostic_Table_Extended",           {})

    accumulate_poll(ALL_DEVICE_DATA)
    logger.info("Named dicts updated from %d device(s).", len(ALL_DEVICE_DATA))


# ===========================================================================
#  HELPER: EXTRACT NUMERIC SERIES FROM A TABLE DICT
# ===========================================================================
def extract_numeric_series(table_dict: Dict[str, Any]) -> Dict[str, Tuple[float, str]]:
    """
    Extract all numeric parameter values from a parsed table dict.

    Parameters
    ----------
    table_dict : Dict[str, Any]
        A dict returned by parse_html_table().

    Returns
    -------
    Dict[str, Tuple[float, str]]
        {param_name: (value, unit)} for every numeric entry.
    """
    series: Dict[str, Tuple[float, str]] = {}
    for key, val in table_dict.items():
        if key.startswith("_") or key.endswith("_value") or key.endswith("_unit"):
            continue
        # key is '#N' → check for corresponding _value
        value_key = f"{key}_value"
        unit_key = f"{key}_unit"
        if value_key in table_dict:
            v = table_dict[value_key]
            u = table_dict.get(unit_key, "")
            if isinstance(v, (int, float)):
                series[val] = (float(v), u)
    return series


# ===========================================================================
#  OUTPUT MODULE 1 — FITS
# ===========================================================================
def write_fits(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
    fits_dir: str,
) -> None:
    """
    Write NRAO-compliant FITS files — one file per device, one BinTableHDU
    per meter table containing the FULL accumulated time-series.

    Layout (each BinTableHDU)
    -------------------------
    Column 0 : TIMESTAMP_LOCAL  — ISO string, format 'A19' (fixed 19-char)
    Column 1+: one 64-bit float column per numeric parameter, all N rows.

    Follows FITS standard (NOST 100-2.0) and NRAO conventions:
      - Primary HDU contains global metadata in header keywords
      - Each table page becomes a FITS BinTableHDU extension
      - Column names truncated to FITS TTYPE limit (68 chars)
      - TELESCOP, INSTRUME, ORIGIN, OBSERVER keywords populated
      - DATE-OBS in ISO-8601 format (first sample timestamp)
      - DATE-END in ISO-8601 format (last sample timestamp)
      - NSAMP keyword: number of accumulated poll cycles
      - BUNIT keyword on each column where units are known
      - All string header values are 7-bit ASCII (NOST 100-2.0 sect. 4.4.2)

    Data is drawn from TIME_SERIES_STORE (full accumulated history), NOT
    from the single-snapshot all_device_data dict.

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: {ip: {table_name: parsed_dict}} — used only for
        iterating device IPs; data comes from TIME_SERIES_STORE.
    fits_dir : str
        Output directory path.
    """
    try:
        from astropy.io import fits as astrofits
        import numpy as np
    except ImportError as exc:
        logger.error("astropy not available — FITS output skipped: %s", exc)
        return

    os.makedirs(fits_dir, exist_ok=True)
    now_utc = datetime.datetime.utcnow()

    # Take a thread-safe snapshot of the full accumulated store
    with _TS_LOCK:
        ts_snapshot = {
            ip: {
                tname: {
                    "timestamps_local": list(tdata["timestamps_local"]),
                    "columns": {p: list(v) for p, v in tdata["columns"].items()},
                    "units":   dict(tdata["units"]),
                }
                for tname, tdata in tables.items()
            }
            for ip, tables in TIME_SERIES_STORE.items()
        }

    for ip in all_device_data:
        safe_ip = ip.replace(".", "_")
        filename = os.path.join(
            fits_dir,
            f"ABMeter_{safe_ip}_{now_utc.strftime('%Y%m%dT%H%M%S')}.fits",
        )

        hdu_list = [astrofits.PrimaryHDU()]
        primary_hdr = hdu_list[0].header

        # --- NRAO / standard FITS primary header keywords ---
        primary_hdr["TELESCOP"] = (_fits_ascii(
            "GBT"),          "Green Bank Telescope facility")
        primary_hdr["INSTRUME"] = (_fits_ascii(
            "ABPowerMeter"), "Allen-Bradley 1403 Site Power Meter")
        primary_hdr["ORIGIN"] = (_fits_ascii(
            "NRAO-GBO"),     "National Radio Astronomy Observatory")
        primary_hdr["OBSERVER"] = (_fits_ascii("WWallace"),     "W. Wallace")
        primary_hdr["DATE-OBS"] = (
            _fits_ascii(now_utc.isoformat(timespec="seconds") + "Z"),
            "UTC file-creation time",
        )
        primary_hdr["FILENAME"] = (_fits_ascii(
            os.path.basename(filename)), "FITS file name")
        primary_hdr["DEVIP"] = (_fits_ascii(ip), "Source device IP address")
        primary_hdr["COMMENT"] = _fits_ascii(
            "Allen-Bradley power meter telemetry - NRAO GBO site infrastructure"
        )
        primary_hdr["HISTORY"] = _fits_ascii(
            f"Generated by ab_power_meter_monitor.py on {now_utc.date()}"
        )

        ip_tables = ts_snapshot.get(ip, {})

        for tname, tdata in ip_tables.items():
            timestamps = tdata["timestamps_local"]
            columns = tdata["columns"]
            units_map = tdata["units"]
            params = list(columns.keys())
            n_samples = len(timestamps)

            if not timestamps:
                logger.debug(
                    "FITS: no accumulated samples for '%s' — skipping HDU", tname)
                continue

            # ----------------------------------------------------------------
            # Column 0: TIMESTAMP_LOCAL — 19-char ASCII strings
            # FITS format 'A19' = fixed-width 19-char ASCII
            # ----------------------------------------------------------------
            fits_cols = [
                astrofits.Column(
                    name=_fits_ascii("TIMESTAMP_LOCAL"),
                    format="A19",
                    unit=_fits_ascii("local time"),
                    array=np.array(timestamps, dtype="U19"),
                )
            ]

            # ----------------------------------------------------------------
            # Columns 1+: one 'D' (float64) column per numeric parameter.
            # None values in the accumulated list are replaced with NaN so
            # the array is densely packed and FITS-compatible.
            # ----------------------------------------------------------------
            for param in params:
                raw_vals = columns[param]
                arr = np.array(
                    [float(v) if v is not None else float("nan")
                     for v in raw_vals],
                    dtype=np.float64,
                )
                col_name = _fits_ascii(param[:68])
                unit_str = _fits_ascii(units_map.get(
                    param, "dimensionless") or "dimensionless")
                fits_cols.append(
                    astrofits.Column(
                        name=col_name,
                        format="D",
                        unit=unit_str,
                        array=arr,
                    )
                )

            hdu = astrofits.BinTableHDU.from_columns(fits_cols)
            ext_name = tname[:8]   # EXTNAME strict 8-char limit
            hdu.header["EXTNAME"] = _fits_ascii(ext_name)
            hdu.header["TBLNAME"] = _fits_ascii(tname)
            hdu.header["SRCIP"] = _fits_ascii(ip)
            hdu.header["NSAMP"] = (
                n_samples, "Number of accumulated poll cycles")
            hdu.header["DATE-OBS"] = _fits_ascii(
                timestamps[0].replace(" ", "T") if timestamps else ""
            )
            hdu.header["DATE-END"] = _fits_ascii(
                timestamps[-1].replace(" ", "T") if timestamps else ""
            )
            hdu.header["COMMENT"] = _fits_ascii(f"AB meter table: {tname}")
            hdu.header["COMMENT"] = _fits_ascii(
                f"{n_samples} sample(s), columnar time-series, TIMESTAMP_LOCAL col 1"
            )
            hdu_list.append(hdu)

        try:
            hdul = astrofits.HDUList(hdu_list)
            hdul.writeto(filename, overwrite=True)
            logger.info("FITS written (%d table HDUs, %s): %s",
                        len(hdu_list) - 1, ip, filename)
        except Exception as exc:
            logger.error("FITS write failed for %s: %s", ip, exc)


# ===========================================================================
#  OUTPUT MODULE 2 — CSV
# ===========================================================================
def write_csv(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
    csv_dir: str,
    append: bool = True,
) -> None:
    """
    Write one CSV file per device per table in COLUMNAR time-series format.

    Layout
    ------
    Row 1 (header) : Timestamp_Local | <param1> | <param2> | ...
                     Units row       : (units)   | <unit1>  | <unit2>  | ...
    Rows 2+        : One row per poll cycle — local timestamp + all values.

    On append (append=True), new rows are added to the existing file.
    The header and units row are written only when the file is created
    for the first time.  Column order is determined on first write and
    preserved; new parameters discovered mid-run are appended as new
    columns on the right with back-filled blanks.

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: {ip: {table_name: parsed_dict}} (used only for
        directory creation; data comes from TIME_SERIES_STORE).
    csv_dir : str
        Output directory path.
    append : bool
        If False, overwrite existing files (start fresh).
    """
    os.makedirs(csv_dir, exist_ok=True)

    with _TS_LOCK:
        snapshot = {
            ip: {
                tname: {
                    "timestamps_local": list(tdata["timestamps_local"]),
                    "columns": {p: list(v) for p, v in tdata["columns"].items()},
                    "units":   dict(tdata["units"]),
                }
                for tname, tdata in tables.items()
            }
            for ip, tables in TIME_SERIES_STORE.items()
        }

    for ip, tables in snapshot.items():
        safe_ip = ip.replace(".", "_")
        for tname, tdata in tables.items():
            filename = os.path.join(csv_dir, f"ABMeter_{safe_ip}_{tname}.csv")
            is_new = not os.path.exists(filename) or not append

            timestamps = tdata["timestamps_local"]
            columns = tdata["columns"]
            units = tdata["units"]
            params = list(columns.keys())

            if not timestamps:
                continue

            try:
                if is_new:
                    # Fresh file: write header, units row, then all accumulated rows
                    with open(filename, "w", newline="", encoding="utf-8") as fh:
                        writer = csv.writer(fh)
                        writer.writerow(["Timestamp_Local"] + params)
                        writer.writerow(
                            ["(units)"] + [units.get(p, "") for p in params])
                        for i, ts in enumerate(timestamps):
                            row = [ts] + [columns[p][i] if i <
                                          len(columns[p]) else "" for p in params]
                            writer.writerow(row)
                else:
                    # Append mode: read existing header to preserve column order,
                    # detect new columns, then append only the new rows.
                    existing_params: List[str] = []
                    existing_row_count = 0
                    with open(filename, "r", newline="", encoding="utf-8") as fh:
                        reader = csv.reader(fh)
                        for row_idx, row in enumerate(reader):
                            if row_idx == 0:
                                # skip Timestamp_Local
                                existing_params = row[1:]
                            elif row_idx == 1:
                                pass  # units row
                            else:
                                existing_row_count += 1

                    new_params = [
                        p for p in params if p not in existing_params]
                    all_params = existing_params + new_params

                    if new_params:
                        # Rewrite file with extended header if new columns appeared
                        with open(filename, "r", newline="", encoding="utf-8") as fh:
                            old_rows = list(csv.reader(fh))
                        with open(filename, "w", newline="", encoding="utf-8") as fh:
                            writer = csv.writer(fh)
                            # Updated header
                            writer.writerow(["Timestamp_Local"] + all_params)
                            # Updated units row
                            all_units = [units.get(p, "") for p in all_params]
                            writer.writerow(["(units)"] + all_units)
                            # Re-emit data rows with blanks for new cols
                            for old_row in old_rows[2:]:
                                writer.writerow(
                                    old_row + [""] * len(new_params))

                    # Append only rows not yet written
                    rows_to_write = timestamps[existing_row_count:]
                    start_idx = existing_row_count
                    with open(filename, "a", newline="", encoding="utf-8") as fh:
                        writer = csv.writer(fh)
                        for i, ts in enumerate(rows_to_write):
                            abs_i = start_idx + i
                            row = [ts] + [
                                columns[p][abs_i] if (
                                    p in columns and abs_i < len(columns[p])) else ""
                                for p in all_params
                            ]
                            writer.writerow(row)

                logger.debug("CSV written (columnar): %s", filename)
            except Exception as exc:
                logger.error("CSV write failed for %s / %s: %s",
                             ip, tname, exc)


# ===========================================================================
#  OUTPUT MODULE 3 — EXCEL (XLSX)
# ===========================================================================
def write_xlsx(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
    xlsx_dir: str,
) -> None:
    """
    Write an Excel workbook per device in COLUMNAR time-series format.

    Layout (each sheet = one table)
    --------------------------------
    Row 1   : Header  — Timestamp_Local | <param1> | <param2> | ...
    Row 2   : Units   — (units)         | <unit1>  | <unit2>  | ...
    Rows 3+ : Data    — local timestamp + values (one row per poll cycle).

    An openpyxl LineChart is appended to each sheet showing all numeric
    parameters as overlaid lines vs. sample index (robust to any N samples).

    The workbook is overwritten each time (XLSX does not support true row
    append without reloading the full file anyway).  All accumulated samples
    from TIME_SERIES_STORE are written.

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: {ip: {table_name: parsed_dict}} (directory only;
        data comes from TIME_SERIES_STORE).
    xlsx_dir : str
        Output directory path.
    """
    try:
        import openpyxl
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        logger.error("openpyxl not available — XLSX output skipped: %s", exc)
        return

    os.makedirs(xlsx_dir, exist_ok=True)

    with _TS_LOCK:
        snapshot = {
            ip: {
                tname: {
                    "timestamps_local": list(tdata["timestamps_local"]),
                    "columns": {p: list(v) for p, v in tdata["columns"].items()},
                    "units":   dict(tdata["units"]),
                }
                for tname, tdata in tables.items()
            }
            for ip, tables in TIME_SERIES_STORE.items()
        }

    for ip, tables in snapshot.items():
        safe_ip = ip.replace(".", "_")
        filename = os.path.join(xlsx_dir, f"ABMeter_{safe_ip}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        header_font = Font(name="Calibri", bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        units_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
        units_font = Font(name="Calibri", bold=False,
                          color="FFFFFF", italic=True)
        header_align = Alignment(horizontal="center")

        for tname, tdata in tables.items():
            timestamps = tdata["timestamps_local"]
            columns = tdata["columns"]
            units_map = tdata["units"]
            params = list(columns.keys())

            if not timestamps:
                continue

            safe_name = tname[:31].replace(
                "/", "_").replace("\\", "_").replace("*", "_")
            ws = wb.create_sheet(title=safe_name)

            # ── Row 1: Headers ──
            header_row = ["Timestamp_Local"] + params
            for col_idx, hdr in enumerate(header_row, start=1):
                cell = ws.cell(row=1, column=col_idx, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # ── Row 2: Units ──
            units_row = ["(units)"] + [units_map.get(p, "") for p in params]
            for col_idx, u in enumerate(units_row, start=1):
                cell = ws.cell(row=2, column=col_idx, value=u)
                cell.font = units_font
                cell.fill = units_fill
                cell.alignment = header_align

            # ── Rows 3+: Data (one row per poll cycle) ──
            for i, ts in enumerate(timestamps):
                data_row_idx = i + 3
                ws.cell(row=data_row_idx, column=1, value=ts)
                for col_idx, param in enumerate(params, start=2):
                    val = columns[param][i] if i < len(
                        columns[param]) else None
                    cell = ws.cell(row=data_row_idx, column=col_idx, value=val)
                    if isinstance(val, (int, float)):
                        cell.number_format = "0.000000"

            n_data_rows = len(timestamps)

            # ── Auto-size columns ──
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            # ── Freeze header + units rows ──
            ws.freeze_panes = "A3"

            # ── Line chart — numeric params vs sample index ──
            numeric_cols = [
                col_idx
                for col_idx, p in enumerate(params, start=2)
                if any(isinstance(v, (int, float)) for v in columns[p] if v is not None)
            ]

            if n_data_rows >= 2 and numeric_cols:
                try:
                    chart = LineChart()
                    chart.title = f"{tname} — {ip}"
                    chart.style = 10
                    chart.y_axis.title = "Value"
                    chart.x_axis.title = "Sample (poll cycle)"
                    chart.width = 24
                    chart.height = 14

                    # cap at 12 series for legibility
                    for nc in numeric_cols[:12]:
                        data_ref = Reference(
                            ws,
                            min_col=nc, max_col=nc,
                            min_row=1, max_row=n_data_rows + 2,
                        )
                        chart.add_data(data_ref, titles_from_data=True)

                    # x-axis categories = Timestamp_Local column
                    cat_ref = Reference(
                        ws, min_col=1, min_row=3, max_row=n_data_rows + 2)
                    chart.set_categories(cat_ref)
                    ws.add_chart(chart, f"A{n_data_rows + 5}")
                except Exception as exc:
                    logger.warning(
                        "Chart creation failed for sheet '%s': %s", safe_name, exc)

        try:
            wb.save(filename)
            logger.info("XLSX written (columnar): %s", filename)
        except Exception as exc:
            logger.error("XLSX save failed for %s: %s", ip, exc)


# ===========================================================================
#  OUTPUT MODULE 4 — TEXT LOG APPEND
# ===========================================================================
def write_log_text(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
    log_dir: str,
) -> None:
    """
    Append new poll rows to per-device, per-table Markdown log files (.md).

    Each file is a valid GitHub-Flavoured Markdown document containing:

    * A level-2 heading with device IP and table name.
    * A GFM pipe table:

      | Timestamp_Local     | Param1 (unit) | Param2 (unit) | ...
      |---------------------|---------------|---------------|----
      | 2026-05-12 08:30:00 | 120.1         | 119.8         | ...
      | 2026-05-12 08:31:00 | 120.3         | 119.6         | ...

    The header and separator rows are written only once when the file is
    created.  Subsequent calls append only new data rows (rows not yet
    persisted), determined by counting existing non-empty, non-separator
    lines after the header.

    Files are named:  ABMeter_<ip>_<table_name>.md

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: {ip: {table_name: parsed_dict}} (directory only;
        data is read from TIME_SERIES_STORE).
    log_dir : str
        Directory in which to create/append Markdown log files.
    """
    os.makedirs(log_dir, exist_ok=True)

    with _TS_LOCK:
        snapshot = {
            ip: {
                tname: {
                    "timestamps_local": list(tdata["timestamps_local"]),
                    "columns": {p: list(v) for p, v in tdata["columns"].items()},
                    "units":   dict(tdata["units"]),
                }
                for tname, tdata in tables.items()
            }
            for ip, tables in TIME_SERIES_STORE.items()
        }

    for ip, tables in snapshot.items():
        safe_ip = ip.replace(".", "_")
        for tname, tdata in tables.items():
            timestamps = tdata["timestamps_local"]
            columns = tdata["columns"]
            units_map = tdata["units"]
            params = list(columns.keys())

            if not timestamps:
                continue

            filename = os.path.join(log_dir, f"ABMeter_{safe_ip}_{tname}.md")

            # ----------------------------------------------------------------
            # Count data rows already persisted.
            # File structure (lines):
            #   1  ## heading
            #   2  blank
            #   3  | header row |
            #   4  | :--- separator |
            #   5+ | data rows |
            # Non-data lines to skip = 4 (heading, blank, header, separator).
            # ----------------------------------------------------------------
            HEADER_LINES = 4   # heading + blank + table header + separator
            existing_rows = 0
            is_new = not os.path.exists(filename)
            if not is_new:
                try:
                    with open(filename, "r", encoding="utf-8") as fh:
                        all_lines = [l for l in fh if l.strip()]
                    # Non-empty lines minus the 3 non-data lines (heading,
                    # header row, separator row — blank line collapses to 0).
                    existing_rows = max(0, len(all_lines) - 3)
                except Exception:
                    existing_rows = 0
                    is_new = True

            # Build column header labels: "Param (unit)" or just "Param"
            col_labels = ["Timestamp_Local"] + [
                f"{p} ({units_map[p]})" if units_map.get(p) else p
                for p in params
            ]

            # Column widths for alignment (min 3 chars for GFM separator).
            # Width = max of header label width and widest data value seen.
            col_widths = [max(3, len(h)) for h in col_labels]
            for i, ts in enumerate(timestamps):
                col_widths[0] = max(col_widths[0], len(ts))
                for j, p in enumerate(params, start=1):
                    val = columns[p][i] if (
                        i < len(columns[p]) and columns[p][i] is not None) else ""
                    col_widths[j] = max(col_widths[j], len(str(val)))

            def _md_row(cells: List[str]) -> str:
                """Format a list of cell strings as a padded GFM table row."""
                padded = [str(c).ljust(col_widths[k])
                          for k, c in enumerate(cells)]
                return "| " + " | ".join(padded) + " |"

            def _md_sep() -> str:
                """GFM left-aligned separator row."""
                return "|" + "|".join("-" * (w + 2) for w in col_widths) + "|"

            try:
                with open(filename, "a", encoding="utf-8") as fh:
                    if is_new:
                        # Heading + blank line + GFM table header + separator
                        fh.write(f"## {ip} — {tname}\n\n")
                        fh.write(_md_row(col_labels) + "\n")
                        fh.write(_md_sep() + "\n")

                    # Append only rows not yet written
                    for i in range(existing_rows, len(timestamps)):
                        ts = timestamps[i]
                        row = [ts] + [
                            str(columns[p][i])
                            if (p in columns and i < len(columns[p]) and columns[p][i] is not None)
                            else ""
                            for p in params
                        ]
                        fh.write(_md_row(row) + "\n")

                logger.debug("Markdown log written: %s", filename)
            except Exception as exc:
                logger.error(
                    "Markdown log write failed for %s / %s: %s", ip, tname, exc)


# ===========================================================================
#  OUTPUT MODULE 5 — VEUSZ  (HDF5 format, Veusz ≥ 3.6 / 4.1)
# ===========================================================================
#
#  Uses the veusz.embed.Embedded API to build the document in-process and
#  saves with mode='hdf5', producing a .vszh5 (HDF5-backed) project file.
#
#  Veusz HDF5 format stores all datasets natively in HDF5 groups, which
#  gives better performance and lossless numeric fidelity compared to the
#  legacy plain-text .vsz format.
#
#  References:
#    • Veusz 3.6 changelog — introduced stable HDF5 save API
#    • Veusz 4.1 — HDF5 is now the recommended/default format
#    • doc.Save(path, mode='hdf5')  — core API call
#    • File extension convention: .vszh5
# ===========================================================================

# Groups of parameter name substrings that share the same SI unit for overlay
# Overlay group display names use spaces (human-readable page/legend titles).
# The keys appear verbatim in Veusz page titles and key widget titles.
VEUSZ_OVERLAY_GROUPS: Dict[str, List[str]] = {
    "Current A":              ["current"],
    "Voltage L-L (V)":        ["l1-l2 voltage", "l2-l3 voltage", "l3-l1 voltage",
                               "3 phase average voltage l-l", "pos. seq. voltage",
                               "neg. seq. voltage", "aux voltage"],
    "Voltage L-N (V)":        ["l1-n voltage", "l2-n voltage", "l3-n voltage",
                               "3 phase average voltage l-n"],
    "Real Power (W)":         ["l1 real power", "l2 real power", "l3 real power",
                               "total real power"],
    "Reactive Power (VAR)":   ["l1 reactive power", "l2 reactive power",
                               "l3 reactive power", "total reactive power"],
    "Apparent Power (VA)":    ["l1 apparent power", "l2 apparent power",
                               "l3 apparent power", "total apparent power"],
    "True PF (%)": ["l1 true pf", "l2 true pf", "l3 true pf",
                    "total true pf"],
    "Displacement PF (%)": ["l1 displacement pf", "l2 displacement pf",
                            "l3 displacement pf", "total displacement pf"],
    "Distortion PF (%)": ["l1 distortion pf", "l2 distortion pf",
                          "l3 distortion pf", "total distortion pf"],
}


def _veusz_safe(name: str) -> str:
    """
    Convert a parameter name to a Veusz-safe dataset identifier.

    Veusz dataset names must not contain spaces, slashes, dots, parentheses
    or other non-word characters.  This function replaces those with
    underscores and strips anything else, returning a string no longer
    than 64 characters.

    Parameters
    ----------
    name : str
        Raw parameter name (e.g. 'L1-L2 Voltage', 'L4(Neutral) Current').

    Returns
    -------
    str
        A valid Veusz dataset name (alphanumerics + underscores only).
    """
    import re
    s = name.replace(" ", "_").replace(".", "_").replace("/", "_")
    s = re.sub(r"[^\w]", "", s)
    return s[:64]


def write_veusz(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
    veusz_dir: str,
    show_window: bool = False,
) -> None:
    """
    Build and save Veusz HDF5 project files (.vszh5) — one per device.

    Uses the **new-style object API** (``embed.Root.Add(...)`` returning
    WidgetNode objects) introduced in Veusz >1.8 and documented at
    https://github.com/veusz/veusz/wiki/EmbeddingPython.  This avoids the
    path-based ``To()`` / ``Set()`` interface entirely, which is fragile and
    requires careful navigation tracking.

    Key API facts confirmed from Veusz source and official wiki:

    * ``graph`` widget has **no** ``title`` setting — use a ``label`` widget
      child to annotate graphs (or encode info in axis labels).
    * ``key`` widget (legend) has a ``title`` setting.
    * Axis settings: ``label`` (string), ``direction`` ('horizontal'/'vertical').
    * xy settings: ``xData``, ``yData``, ``marker``, ``key`` (legend text),
      ``PlotLine/width`` (via the node: ``xy.PlotLine.width.val = '1.5pt'``).
    * Settings are read/written via ``.val`` on SettingNode objects, e.g.
      ``axis.label.val = 'Voltage [V]'``.
    * ``doc.Save(path, mode='hdf5')`` — saves as HDF5 (.vszh5), Veusz >= 3.6.

    Parameters
    ----------
    all_device_data : Dict[str, Dict[str, Dict[str, Any]]]
        Nested dict: ``{ip: {table_name: parsed_dict}}``.
    veusz_dir : str
        Output directory path (created if absent).
    show_window : bool, optional
        If True the embedded Veusz window is shown (toolbar visible for
        interactive inspection).  Defaults to False (headless / silent).

    Raises
    ------
    ImportError
        Logged and skipped if ``veusz`` is not installed.
        Install with: ``pip install veusz``.
    """
    try:
        import veusz.embed as vz
    except ImportError as exc:
        logger.error(
            "veusz package not available — Veusz HDF5 output skipped: %s\n"
            "Install with: pip install veusz",
            exc,
        )
        return

    os.makedirs(veusz_dir, exist_ok=True)
    now_utc = datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"

    for ip, tables in all_device_data.items():
        safe_ip = ip.replace(".", "_")
        filename = os.path.join(veusz_dir, f"ABMeter_{safe_ip}.vszh5")

        # -------------------------------------------------------------------
        # Open the embedded document window.
        # hidden=True  → no GUI window (headless/background).
        # hidden=False → opens the full Veusz application window with toolbar.
        # -------------------------------------------------------------------
        win_title = f"AB Power Meter - {ip} - {now_utc}"   # ASCII only
        try:
            doc = vz.Embedded(win_title, hidden=not show_window)
        except Exception as exc:
            logger.error("Failed to open Veusz Embedded for %s: %s", ip, exc)
            continue

        try:
            # ---------------------------------------------------------------
            # Step 1 — Collect all numeric series and load datasets.
            #
            # Each numeric parameter becomes a 1-D dataset named
            #   <table_safe>_<param_safe>
            # A companion index dataset (idx_<name>) provides the x-axis
            # so that future poll snapshots can be appended.
            # ---------------------------------------------------------------
            all_series: Dict[str, Dict[str, Tuple[float, str]]] = {}

            for tname, tdict in tables.items():
                series = extract_numeric_series(tdict)
                if not series:
                    logger.debug(
                        "Veusz: no numeric data in '%s' — skipping", tname)
                    continue
                all_series[tname] = series

            if not all_series:
                logger.warning(
                    "Veusz: no numeric data for device %s — skipping", ip)
                doc.Close()
                continue

            # ---------------------------------------------------------------
            # Load ALL accumulated time-series data from TIME_SERIES_STORE.
            # Each dataset is a list of all values across every poll cycle.
            # The x-axis dataset is the sample index (0, 1, 2, …) so that
            # Veusz plots each parameter as a line over time.
            # ---------------------------------------------------------------
            with _TS_LOCK:
                ts_ip = TIME_SERIES_STORE.get(ip, {})

            n_datasets = 0
            for tname, series in all_series.items():
                tstore = ts_ip.get(tname, {})
                ts_columns = tstore.get("columns", {})
                timestamps = tstore.get("timestamps_local", [])
                n_samples = len(timestamps)

                for param in series:
                    ds_name = _veusz_safe(f"{tname}_{param}")
                    idx_name = _veusz_safe(f"idx_{tname}_{param}")

                    if param in ts_columns and ts_columns[param]:
                        # Use full accumulated series; replace None with NaN
                        raw = ts_columns[param]
                        vals = [float(v) if v is not None else float("nan")
                                for v in raw]
                        idxs = [float(k) for k in range(len(vals))]
                    else:
                        # Fallback: single-point from latest snapshot
                        vals = [float(series[param][0])]
                        idxs = [0.0]

                    doc.SetData(ds_name,  vals)
                    doc.SetData(idx_name, idxs)
                    n_datasets += 1

            logger.debug("Veusz: loaded %d datasets (%d samples each) for %s",
                         n_datasets, n_samples if n_samples else 1, ip)

            # ---------------------------------------------------------------
            # Step 2 — Per-table pages (new-style object API).
            #
            # Widget tree per page:
            #   page
            #     grid  (2-column layout)
            #       graph  (one per numeric parameter, autoadd=False)
            #         axis 'x'  (horizontal, label='Sample Index')
            #         axis 'y'  (vertical,   label='<param> [<unit>]')
            #         xy        (xData=idx_ds, yData=val_ds)
            #         label     (annotation text — graph has no title setting)
            #
            # IMPORTANT: graph widget has NO 'title' property.  The correct
            # way to annotate a graph is to add a 'label' widget child and
            # set its 'label' setting (the text string), position it at the
            # top-centre using xPos/yPos fractional coordinates.
            # ---------------------------------------------------------------
            root = doc.Root   # WidgetNode for the document root

            # ---------------------------------------------------------------
            # Colour palette — large distinguishable set for automatic cycling.
            # These are standard CSS/SVG colour names that Veusz recognises.
            # A modulo index is applied so any number of series gets a colour.
            # ---------------------------------------------------------------
            COLOUR_CYCLE = [
                "red",        "blue",       "green",      "darkorange",
                "purple",     "deeppink",   "teal",       "saddlebrown",
                "navy",       "olive",      "crimson",    "darkgreen",
                "royalblue",  "darkorchid", "darkcyan",   "chocolate",
                "darkred",    "steelblue",  "seagreen",   "goldenrod",
                "indigo",     "mediumblue", "firebrick",  "darkslategray",
            ]

            def _colour(idx: int) -> str:
                """Return a colour name from COLOUR_CYCLE by cyclic index."""
                return COLOUR_CYCLE[idx % len(COLOUR_CYCLE)]

            def _human(name: str) -> str:
                """Replace underscores with spaces for display labels."""
                return name.replace("_", " ")

            for tname, series in all_series.items():

                # --- Page — human-readable name (spaces, no underscores) ---
                page_wname = _veusz_safe(f"{tname}_page")
                page = root.Add("page", name=page_wname, autoadd=False)

                # --- Grid (2 columns) ---
                grid = page.Add("grid", name="grid1", autoadd=False)
                grid.rows.val = max(1, (len(series) + 1) // 2)
                grid.columns.val = 2

                for p_idx, (param, (val, unit)) in enumerate(series.items()):
                    ds_name = _veusz_safe(f"{tname}_{param}")
                    idx_name = _veusz_safe(f"idx_{tname}_{param}")
                    gname = _veusz_safe(f"g_{param}")
                    colour = _colour(p_idx)
                    # Y-axis label: human-readable param + unit, no underscores
                    axis_label = f"{_human(param)} [{unit}]" if unit else _human(
                        param)

                    # --- Graph ---
                    graph = grid.Add("graph", name=gname, autoadd=False)

                    # --- x-axis ---
                    ax = graph.Add("axis", name="x", autoadd=False)
                    ax.label.val = "Sample Index (poll cycle)"
                    ax.direction.val = "horizontal"

                    # --- y-axis (carries the parameter label) ---
                    ay = graph.Add("axis", name="y", autoadd=False)
                    ay.label.val = axis_label
                    ay.direction.val = "vertical"

                    # --- xy plotter with auto colour ---
                    xy = graph.Add("xy", name="plot1", autoadd=False)
                    xy.xData.val = idx_name
                    xy.yData.val = ds_name
                    xy.marker.val = "circle"
                    xy.PlotLine.width.val = "1.5pt"
                    # Line and marker colour — set the same colour on both so
                    # line and marker are always consistent.
                    try:
                        xy.PlotLine.color.val = colour
                        xy.MarkerFill.color.val = colour
                        xy.MarkerLine.color.val = colour
                    except Exception:
                        pass   # older Veusz versions may use different attr names

                    # --- label widget for graph title annotation ---
                    # graph has no 'title' property; use a label widget instead.
                    lbl = graph.Add("label", name="title lbl", autoadd=False)
                    lbl.label.val = f"{_human(param)} ({ip})"
                    lbl.xPos.val = [0.5]   # horizontally centred
                    lbl.yPos.val = [1.02]  # just above the plot area

            # ---------------------------------------------------------------
            # Step 3 — Overlay pages.
            # One page per unit group; all parameters sharing the same unit
            # are overlaid on a single graph.  Each xy series gets a unique
            # colour from COLOUR_CYCLE.  A 'key' widget provides the legend.
            # ---------------------------------------------------------------
            for group_label, substrings in VEUSZ_OVERLAY_GROUPS.items():

                # Collect (ds_name, idx_name, param, unit) tuples
                overlay: List[Tuple[str, str, str, str]] = []
                for tname, series in all_series.items():
                    for param, (val, unit) in series.items():
                        if any(sub in param.lower() for sub in substrings):
                            ds_name = _veusz_safe(f"{tname}_{param}")
                            idx_name = _veusz_safe(f"idx_{tname}_{param}")
                            overlay.append((ds_name, idx_name, param, unit))

                if not overlay:
                    continue

                first_unit = overlay[0][3] if overlay else ""
                y_label = group_label   # already human-readable (spaces)
                if first_unit:
                    y_label = f"{group_label} [{first_unit}]"

                # --- Page — human-readable, no underscores ---
                ov_page_name = _veusz_safe(f"overlay_{group_label}")
                ov_page = root.Add("page", name=ov_page_name, autoadd=False)

                # --- Single graph ---
                ov_graph = ov_page.Add(
                    "graph", name="overlay graph", autoadd=False)

                # --- Axes ---
                ox = ov_graph.Add("axis", name="x", autoadd=False)
                ox.label.val = "Sample Index (poll cycle)"
                ox.direction.val = "horizontal"

                oy = ov_graph.Add("axis", name="y", autoadd=False)
                oy.label.val = y_label
                oy.direction.val = "vertical"

                # --- Key / legend ---
                key_wgt = ov_graph.Add("key", name="key1", autoadd=False)
                key_wgt.title.val = f"Overlay: {group_label}"

                # --- One xy per overlaid parameter, each with a unique colour ---
                for ov_idx, (ds_name, idx_name, param, unit) in enumerate(overlay):
                    xy_wname = _veusz_safe(f"xy_{ds_name}")
                    colour = _colour(ov_idx)
                    xy = ov_graph.Add("xy", name=xy_wname, autoadd=False)
                    xy.xData.val = idx_name
                    xy.yData.val = ds_name
                    # legend entry (no underscores)
                    xy.key.val = _human(param)
                    xy.marker.val = "circle"
                    xy.PlotLine.width.val = "1.5pt"
                    try:
                        xy.PlotLine.color.val = colour
                        xy.MarkerFill.color.val = colour
                        xy.MarkerLine.color.val = colour
                    except Exception:
                        pass

                # --- label widget for overlay page title ---
                ov_lbl = ov_graph.Add("label", name="title lbl", autoadd=False)
                ov_lbl.label.val = f"Overlay: {group_label} - {ip}"
                ov_lbl.xPos.val = [0.5]
                ov_lbl.yPos.val = [1.02]

            # ---------------------------------------------------------------
            # Step 4 — Save as HDF5 (.vszh5).
            # mode='hdf5' is supported from Veusz 3.6 and is the recommended
            # default in Veusz 4.1.  All datasets are stored in native HDF5
            # groups for compact, lossless, random-access storage.
            # ---------------------------------------------------------------
            doc.Save(filename, mode="hdf5")
            logger.info("Veusz HDF5 project saved: %s", filename)

        except Exception as exc:
            logger.error(
                "Veusz build/save failed for device %s: %s\n%s",
                ip, exc, traceback.format_exc(),
            )
        finally:
            # Always close to release the Veusz process resources, even on
            # error.  Harmless if the doc was already closed.
            try:
                doc.Close()
            except Exception:
                pass


def open_veusz_preview(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
    veusz_dir: str,
) -> None:
    """
    Convenience wrapper: call write_veusz with show_window=True so the
    Veusz application window opens with the toolbar visible for interactive
    inspection, then save to .vszh5 on close.

    This is the function wired to the "Open in Veusz" button in the GUI.

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: ``{ip: {table_name: parsed_dict}}``.
    veusz_dir : str
        Output directory path.
    """
    write_veusz(all_device_data, veusz_dir, show_window=True)


def flush_outputs_parallel(cfg: Dict[str, Any]) -> "concurrent.futures.Future":
    """
    Dispatch non-Veusz output writers to a thread-pool executor so that
    CSV / XLSX / log writes happen concurrently with ongoing polling.

    Veusz is deliberately excluded — it is always built once at loop end
    from the complete accumulated dataset.

    The caller receives a Future object; the flush completes asynchronously.
    The caller should NOT await it synchronously in the poll thread — just
    fire and forget.  If the previous flush is still running when a new one
    is triggered, the new one is skipped to prevent file corruption.

    Parameters
    ----------
    cfg : Dict[str, Any]
        Runtime configuration dict.

    Returns
    -------
    concurrent.futures.Future
        Future representing the background flush task.
    """
    def _do_flush() -> None:
        fits_dir = cfg.get("fits_dir",  FITS_DIR)
        csv_dir = cfg.get("csv_dir",   CSV_DIR)
        xlsx_dir = cfg.get("xlsx_dir",  XLSX_DIR)
        log_dir = cfg.get("log_dir",   LOG_DIR)

        tasks = []
        if cfg.get("enable_fits"):
            tasks.append(
                ("FITS", lambda: write_fits(ALL_DEVICE_DATA, fits_dir)))
        if cfg.get("enable_csv"):
            tasks.append(("CSV", lambda: write_csv(
                ALL_DEVICE_DATA, csv_dir, append=True)))
        if cfg.get("enable_xlsx"):
            tasks.append(
                ("XLSX", lambda: write_xlsx(ALL_DEVICE_DATA, xlsx_dir)))
        if cfg.get("enable_log_append"):
            tasks.append(
                ("LOG", lambda: write_log_text(ALL_DEVICE_DATA, log_dir)))

        with concurrent.futures.ThreadPoolExecutor(
            max_workers=min(4, len(tasks)) if tasks else 1,
            thread_name_prefix="ab_flush",
        ) as ex:
            futs = {ex.submit(fn): name for name, fn in tasks}
            for fut in concurrent.futures.as_completed(futs):
                name = futs[fut]
                try:
                    fut.result()
                    logger.debug("Parallel flush OK: %s", name)
                except Exception as exc:
                    logger.error("Parallel flush ERROR (%s): %s", name, exc)

    # Use a module-level executor so we can check if one is already running
    ex = concurrent.futures.ThreadPoolExecutor(
        max_workers=1, thread_name_prefix="ab_flush_mgr")
    future = ex.submit(_do_flush)
    ex.shutdown(wait=False)
    return future


# ===========================================================================
#  MATPLOTLIB PREVIEW HELPER (used by GUI)
# ===========================================================================
def build_preview_figures(
    all_device_data: Dict[str, Dict[str, Dict[str, Any]]],
) -> List[Any]:
    """
    Build a list of matplotlib Figure objects for GUI preview display.

    Creates:
      • One figure per table (per device) showing all numeric params as a bar chart.
      • One overlay figure per unit-group containing data across tables.

    Parameters
    ----------
    all_device_data : Dict
        Nested dict: {ip: {table_name: parsed_dict}}

    Returns
    -------
    List[matplotlib.figure.Figure]
        List of Figure objects ready for embedding in a Qt canvas.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")   # non-interactive backend for embedding
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker
    except ImportError as exc:
        logger.error("matplotlib not available — preview skipped: %s", exc)
        return []

    # Close any previously opened figures to prevent the matplotlib warning
    # "More than 20 figures have been opened" — each call to this function
    # replaces the prior figure set, so old ones must be explicitly closed.
    plt.close("all")

    figures: List[Any] = []
    prop_cycle_colors = plt.rcParams["axes.prop_cycle"].by_key()["color"]

    for ip, tables in all_device_data.items():
        # --- Per-table figures ---
        for tname, tdict in tables.items():
            series = extract_numeric_series(tdict)
            if not series:
                continue

            params = list(series.keys())
            values = [series[p][0] for p in params]
            units = [series[p][1] for p in params]

            fig, ax = plt.subplots(figsize=(10, 4))
            x_pos = range(len(params))
            bars = ax.bar(
                x_pos, values, color=prop_cycle_colors[:len(params)] * 10)
            ax.set_xticks(list(x_pos))
            ax.set_xticklabels(params, rotation=45, ha="right", fontsize=7)
            ax.set_title(f"{tname}\n{ip}", fontsize=9)
            ax.set_ylabel("Value")
            ax.grid(axis="y", alpha=0.3)
            fig.tight_layout()
            fig._ab_title = f"{ip} — {tname}"  # type: ignore[attr-defined]
            figures.append(fig)

        # --- Overlay figures by unit group ---
        for group_label, substrings in VEUSZ_OVERLAY_GROUPS.items():
            group_params: List[str] = []
            group_values: List[float] = []
            group_labels: List[str] = []

            for tname, tdict in tables.items():
                series = extract_numeric_series(tdict)
                for param, (val, unit) in series.items():
                    if any(s in param.lower() for s in substrings):
                        group_params.append(param)
                        group_values.append(val)
                        group_labels.append(f"{tname[:12]}\n{param[:20]}")

            if len(group_values) < 2:
                continue

            fig, ax = plt.subplots(figsize=(10, 4))
            x_pos = range(len(group_labels))
            ax.bar(x_pos, group_values,
                   color=prop_cycle_colors[:len(group_labels)] * 10)
            ax.set_xticks(list(x_pos))
            ax.set_xticklabels(group_labels, rotation=45,
                               ha="right", fontsize=7)
            ax.set_title(f"Overlay: {group_label}\n{ip}", fontsize=9)
            ax.set_ylabel(group_label)
            ax.grid(axis="y", alpha=0.3)
            fig.tight_layout()
            # type: ignore[attr-defined]
            fig._ab_title = f"{ip} — Overlay: {group_label}"
            figures.append(fig)

    return figures


# ===========================================================================
#  GUI — PyQt (PySide6 via QtPy abstraction)
# ===========================================================================
# QtPy transparently wraps PySide6 (or PyQt6 as fallback).
# Set QT_API env var to force one: export QT_API=pyside6


class _QTextEditHandler(logging.Handler):
    """
    A ``logging.Handler`` that appends formatted log records to a
    ``QTextEdit`` widget in real time.

    This is the bridge that makes every ``logger.*()`` call — including
    connection warnings from ``fetch_table_html()``, parse errors,
    FITS/CSV/XLSX failures, and Veusz messages — appear live in the
    GUI status console without any extra ``_append_log()`` calls scattered
    through the code.

    Usage
    -----
    Instantiate once, pass the target QTextEdit, then add to the module
    logger::

        handler = _QTextEditHandler(self._log_console)
        logging.getLogger("ABMonitor").addHandler(handler)

    Remove on window close to avoid writing to a destroyed widget::

        logging.getLogger("ABMonitor").removeHandler(handler)

    Thread safety
    -------------
    ``emit()`` uses ``QMetaObject.invokeMethod`` with
    ``Qt.ConnectionType.QueuedConnection`` so records originating on the
    background ``PollThread`` are safely marshalled to the GUI thread
    before touching the widget.
    """

    # Colour map: log level -> HTML colour for the console text
    _LEVEL_COLOUR: Dict[int, str] = {
        logging.DEBUG:    "#6c7086",   # muted grey
        logging.INFO:     "#cdd6f4",   # default text
        logging.WARNING:  "#f9e2af",   # yellow
        logging.ERROR:    "#f38ba8",   # red
        logging.CRITICAL: "#ff5555",   # bright red
    }

    def __init__(self, widget: Any, level: int = logging.DEBUG) -> None:
        """
        Parameters
        ----------
        widget : QTextEdit
            The console widget to append records to.
        level : int
            Minimum logging level to display (default DEBUG — show all).
        """
        super().__init__(level)
        self._widget = widget
        self.setFormatter(
            logging.Formatter(
                fmt="%(asctime)s  %(levelname)-8s  %(message)s",
                datefmt="%H:%M:%S",
            )
        )

    def emit(self, record: logging.LogRecord) -> None:
        """
        Append a formatted, coloured log line to the QTextEdit.

        Called by the logging framework on every matching record.
        Uses a queued cross-thread invoke so it is safe from any thread.

        Parameters
        ----------
        record : logging.LogRecord
            The log record to display.
        """
        try:
            msg = self.format(record)
            colour = self._LEVEL_COLOUR.get(record.levelno, "#cdd6f4")
            # Escape HTML special chars so angle brackets in messages render
            # correctly rather than being interpreted as HTML tags.
            escaped = (
                msg.replace("&", "&amp;")
                   .replace("<", "&lt;")
                   .replace(">", "&gt;")
            )
            html = f'<span style="color:{colour};">{escaped}</span>'

            # Marshal to the GUI thread via a queued invoke.
            # This is safe whether emit() is called from the main thread
            # or from PollThread.
            try:
                from qtpy.QtCore import QMetaObject, Qt
                from qtpy.QtCore import Q_ARG
                QMetaObject.invokeMethod(
                    self._widget,
                    "append",
                    Qt.ConnectionType.QueuedConnection,
                    Q_ARG(str, html),
                )
            except Exception:
                # Fallback: direct call (only safe on GUI thread)
                self._widget.append(html)

            # Auto-scroll to bottom
            try:
                from qtpy.QtCore import QMetaObject, Qt
                QMetaObject.invokeMethod(
                    self._widget.verticalScrollBar(),
                    "setValue",
                    Qt.ConnectionType.QueuedConnection,
                    Q_ARG(int, self._widget.verticalScrollBar().maximum()),
                )
            except Exception:
                pass

        except Exception:
            # Never let a logging handler crash the application
            self.handleError(record)


def launch_gui(
    initial_switches: Dict[str, Any],
    initial_figures:  List[Any],
) -> None:
    """
    Launch the main PyQt/PySide6 GUI window.

    The GUI provides:
      • Light / dark theme toggle (menu)
      • Check-boxes for each output switch
      • IP range last-octet spin-boxes
      • Sample period spin-box
      • Log directory file chooser
      • Live plot preview (matplotlib FigureCanvas)
      • Poll Now / Start / Stop buttons

    Parameters
    ----------
    initial_switches : Dict[str, Any]
        Dict of switch states loaded from module-level config variables.
    initial_figures : List
        Pre-computed matplotlib Figure objects for initial display.
    """
    import os
    os.environ.setdefault("QT_API", "pyside6")

    try:
        from qtpy import QtWidgets, QtCore, QtGui
        from qtpy.QtWidgets import (
            QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
            QGridLayout, QGroupBox, QCheckBox, QSpinBox, QDoubleSpinBox,
            QLabel, QPushButton, QFileDialog, QLineEdit, QTabWidget,
            QScrollArea, QSizePolicy, QMenuBar, QMenu, QAction, QStatusBar,
            QTextEdit, QSplitter,
        )
        from qtpy.QtCore import Qt, QTimer, Signal, QThread
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavToolbar
    except ImportError as exc:
        logger.critical(
            "GUI dependencies missing: %s\nInstall: pip install qtpy pyside6 matplotlib", exc)
        return

    # -----------------------------------------------------------------------
    # Background polling thread
    # -----------------------------------------------------------------------
    class PollThread(QThread):
        """Worker thread that polls devices on a configurable interval."""

        data_ready = Signal(dict)    # emits all_device_data dict each cycle
        error_occur = Signal(str)     # emits error description string
        log_message = Signal(str)     # emits log text for status console

        def __init__(self, config: Dict[str, Any], parent=None):
            super().__init__(parent)
            self.config = config
            self._running = False

        def run(self) -> None:
            self._running = True
            while self._running:
                try:
                    self.log_message.emit(
                        f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Polling …"
                    )
                    data = poll_all_devices(
                        ip_base=self.config["ip_base"],
                        octet_start=self.config["octet_start"],
                        octet_end=self.config["octet_end"],
                        table_names=TABLE_NAMES,
                    )
                    update_named_dicts(data)
                    self.data_ready.emit(data)
                    self.log_message.emit(
                        f"[{datetime.datetime.now().strftime('%H:%M:%S')}] "
                        f"Poll complete — {len(data)} table dicts."
                    )
                except Exception as exc:
                    self.error_occur.emit(f"Poll error: {exc}")

                # Sleep in 0.5 s chunks so stop() is responsive
                remaining = self.config.get("sample_period", SAMPLE_PERIOD_SEC)
                while remaining > 0 and self._running:
                    time.sleep(min(0.5, remaining))
                    remaining -= 0.5

        def stop(self) -> None:
            self._running = False

    # -----------------------------------------------------------------------
    # Main Window
    # -----------------------------------------------------------------------
    class MainWindow(QMainWindow):
        """
        Primary application window for the AB Power Meter Monitor.
        """

        LIGHT_STYLE = ""   # use Qt default

        DARK_STYLE = """
            QMainWindow, QWidget, QDialog {
                background-color: #1e1e2e;
                color: #cdd6f4;
            }
            QGroupBox {
                background-color: #181825;
                border: 1px solid #45475a;
                border-radius: 6px;
                margin-top: 10px;
                padding: 8px;
                color: #89b4fa;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
            }
            QCheckBox, QLabel, QSpinBox, QDoubleSpinBox, QLineEdit {
                color: #cdd6f4;
                background-color: transparent;
            }
            QPushButton {
                background-color: #313244;
                color: #cdd6f4;
                border: 1px solid #45475a;
                border-radius: 4px;
                padding: 4px 12px;
            }
            QPushButton:hover  { background-color: #45475a; }
            QPushButton:pressed{ background-color: #585b70; }
            QSpinBox, QDoubleSpinBox, QLineEdit {
                background-color: #313244;
                border: 1px solid #45475a;
                border-radius: 3px;
                padding: 2px;
            }
            QTabWidget::pane { border: 1px solid #45475a; }
            QTabBar::tab {
                background: #313244;
                color: #cdd6f4;
                padding: 6px 16px;
            }
            QTabBar::tab:selected { background: #45475a; color: #89b4fa; }
            QTextEdit {
                background-color: #181825;
                color: #a6e3a1;
                border: 1px solid #45475a;
                font-family: "Courier New", monospace;
                font-size: 10pt;
            }
            QScrollBar:vertical {
                background: #1e1e2e;
                width: 10px;
            }
            QScrollBar::handle:vertical {
                background: #45475a;
                border-radius: 5px;
            }
            QStatusBar { background: #181825; color: #a6adc8; }
            QMenuBar { background: #181825; color: #cdd6f4; }
            QMenuBar::item:selected { background: #313244; }
            QMenu { background: #1e1e2e; color: #cdd6f4; border: 1px solid #45475a; }
            QMenu::item:selected { background: #313244; }
        """

        def __init__(self, switches: Dict[str, Any], figures: List[Any]) -> None:
            super().__init__()
            self.setWindowTitle("AB Power Meter Monitor — NRAO / GBO")
            self.resize(1280, 820)

            self._switches = dict(switches)
            self._figures = list(figures)
            self._thread: Optional[PollThread] = None
            self._dark_mode = False
            # attached after widget is built
            self._log_handler: Optional[logging.Handler] = None
            # Tracks in-flight background file flush so we never start two at once
            self._flush_future: Optional["concurrent.futures.Future"] = None

            self._build_menu()
            self._build_central()      # builds self._log_console
            self._build_status_bar()

            # --- Attach the QTextEditHandler to the module logger so that
            # every logger.info/warning/error call in the entire codebase
            # (poll errors, FITS issues, CSV writes, Veusz, etc.) appears
            # live in the status console with colour coding by severity.
            self._log_handler = _QTextEditHandler(
                self._log_console, level=logging.DEBUG)
            logging.getLogger("ABMonitor").addHandler(self._log_handler)

            # Apply initial dark mode if OS prefers it
            if QtWidgets.QApplication.instance().palette().window().color().lightness() < 128:
                self._apply_dark()

        # ----------------------------------------------------------------
        # Menu bar
        # ----------------------------------------------------------------
        def _build_menu(self) -> None:
            menubar = self.menuBar()

            # View menu — theme toggle
            view_menu = menubar.addMenu("&View")
            self._act_toggle_theme = QAction("Switch to &Dark Theme", self)
            self._act_toggle_theme.triggered.connect(self._toggle_theme)
            view_menu.addAction(self._act_toggle_theme)

            # File menu
            file_menu = menubar.addMenu("&File")
            act_quit = QAction("&Quit", self)
            act_quit.triggered.connect(self.close)
            file_menu.addAction(act_quit)

            # Help menu
            help_menu = menubar.addMenu("&Help")
            act_about = QAction("&About", self)
            act_about.triggered.connect(self._show_about)
            help_menu.addAction(act_about)

        # ----------------------------------------------------------------
        # Central widget
        # ----------------------------------------------------------------
        def _build_central(self) -> None:
            central = QWidget()
            main_layout = QVBoxLayout(central)
            main_layout.setSpacing(6)

            splitter = QSplitter(Qt.Orientation.Horizontal)

            # ---- Left panel: controls ----
            ctrl_widget = QWidget()
            ctrl_layout = QVBoxLayout(ctrl_widget)
            ctrl_layout.setSpacing(6)

            ctrl_layout.addWidget(self._build_ip_group())
            ctrl_layout.addWidget(self._build_timing_group())
            ctrl_layout.addWidget(self._build_output_group())
            ctrl_layout.addWidget(self._build_action_buttons())
            ctrl_layout.addStretch()

            self._log_console = QTextEdit()
            self._log_console.setReadOnly(True)
            self._log_console.setMaximumHeight(180)
            self._log_console.setPlaceholderText("Status / log output …")
            ctrl_layout.addWidget(QLabel("Status Console"))
            ctrl_layout.addWidget(self._log_console)

            ctrl_widget.setMaximumWidth(340)
            splitter.addWidget(ctrl_widget)

            # ---- Right panel: plot tabs ----
            self._tab_widget = QTabWidget()
            self._populate_plot_tabs(self._figures)
            splitter.addWidget(self._tab_widget)
            splitter.setStretchFactor(1, 1)

            main_layout.addWidget(splitter)
            self.setCentralWidget(central)

        def _build_ip_group(self) -> QGroupBox:
            """Build the IP address range control group."""
            grp = QGroupBox("IP Address Range  (10.16.130.X)")
            layout = QGridLayout()

            layout.addWidget(QLabel("Last Octet Start:"), 0, 0)
            self._spin_ip_start = QSpinBox()
            self._spin_ip_start.setRange(1, 254)
            self._spin_ip_start.setValue(self._switches.get(
                "octet_start", IP_LAST_OCTET_START))
            layout.addWidget(self._spin_ip_start, 0, 1)

            layout.addWidget(QLabel("Last Octet End:"), 1, 0)
            self._spin_ip_end = QSpinBox()
            self._spin_ip_end.setRange(1, 254)
            self._spin_ip_end.setValue(
                self._switches.get("octet_end", IP_LAST_OCTET_END))
            layout.addWidget(self._spin_ip_end, 1, 1)

            grp.setLayout(layout)
            return grp

        def _build_timing_group(self) -> QGroupBox:
            """Build the sample period control group."""
            grp = QGroupBox("Timing")
            layout = QGridLayout()

            layout.addWidget(QLabel("Sample Period (s):"), 0, 0)
            self._spin_period = QDoubleSpinBox()
            self._spin_period.setRange(5.0, 3600.0)
            self._spin_period.setSingleStep(5.0)
            self._spin_period.setDecimals(1)
            self._spin_period.setValue(self._switches.get(
                "sample_period", SAMPLE_PERIOD_SEC))
            layout.addWidget(self._spin_period, 0, 1)

            grp.setLayout(layout)
            return grp

        def _build_output_group(self) -> QGroupBox:
            """Build the output enable check-boxes and log-dir chooser."""
            grp = QGroupBox("Output Options")
            layout = QVBoxLayout()

            # Check-boxes for the 5 file-output modes.
            # NOTE: "Enable GUI" is intentionally omitted here — the GUI is
            # already running, so that switch is only meaningful in the
            # ENABLE_GUI header variable and has no in-app toggle.
            self._cb_fits = QCheckBox("Enable FITS output")
            self._cb_csv = QCheckBox("Enable CSV output")
            self._cb_xlsx = QCheckBox("Enable Excel (XLSX) output")
            self._cb_log = QCheckBox("Append to log files")
            self._cb_veusz = QCheckBox("Enable Veusz HDF5 output (.vszh5)")

            self._cb_fits.setChecked(
                bool(self._switches.get("enable_fits",       ENABLE_FITS)))
            self._cb_csv.setChecked(
                bool(self._switches.get("enable_csv",        ENABLE_CSV)))
            self._cb_xlsx.setChecked(
                bool(self._switches.get("enable_xlsx",       ENABLE_XLSX)))
            self._cb_log.setChecked(bool(self._switches.get(
                "enable_log_append", ENABLE_LOG_APPEND)))
            self._cb_veusz.setChecked(
                bool(self._switches.get("enable_veusz",    ENABLE_VEUSZ)))

            for cb in [self._cb_fits, self._cb_csv, self._cb_xlsx, self._cb_log, self._cb_veusz]:
                layout.addWidget(cb)

            # --- Output root directory (drives FITS, CSV, XLSX, Veusz sub-dirs) ---
            layout.addWidget(QLabel("Output Root Directory:"))
            out_dir_layout = QHBoxLayout()
            self._le_out_dir = QLineEdit(self._switches.get(
                "output_base_dir", OUTPUT_BASE_DIR))
            self._le_out_dir.setPlaceholderText(
                "Base directory for all output files …")
            self._le_out_dir.setToolTip(
                "All sub-directories (fits/, csv/, xlsx/, veusz/, logs/) are "
                "created inside this folder.  Mirrors the OUTPUT_BASE_DIR "
                "header variable."
            )
            btn_browse_out = QPushButton("Browse …")
            btn_browse_out.clicked.connect(self._choose_out_dir)
            out_dir_layout.addWidget(self._le_out_dir)
            out_dir_layout.addWidget(btn_browse_out)
            layout.addLayout(out_dir_layout)

            # --- Log directory override (defaults to <output_root>/logs) ---
            layout.addWidget(QLabel("Log File Directory (override):"))
            log_dir_layout = QHBoxLayout()
            self._le_log_dir = QLineEdit(
                self._switches.get("log_dir", LOG_DIR))
            self._le_log_dir.setPlaceholderText(
                "Log file directory (leave blank to use output root) …")
            self._le_log_dir.setToolTip(
                "Override only the log directory.  Leave blank to use "
                "<Output Root>/logs/ automatically."
            )
            btn_browse_log = QPushButton("Browse …")
            btn_browse_log.clicked.connect(self._choose_log_dir)
            log_dir_layout.addWidget(self._le_log_dir)
            log_dir_layout.addWidget(btn_browse_log)
            layout.addLayout(log_dir_layout)

            grp.setLayout(layout)
            return grp

        def _build_action_buttons(self) -> QWidget:
            """Build Poll Now / Start / Stop action buttons."""
            widget = QWidget()
            layout = QHBoxLayout(widget)

            self._btn_poll = QPushButton("Poll Now")
            self._btn_start = QPushButton("Start Auto")
            self._btn_stop = QPushButton("Stop")
            self._btn_veusz = QPushButton("Open in Veusz")
            self._btn_stop.setEnabled(False)
            self._btn_veusz.setToolTip(
                "Build plots in the live Veusz window and save as .vszh5 (HDF5)"
            )

            self._btn_poll.clicked.connect(self._do_poll_once)
            self._btn_start.clicked.connect(self._do_start)
            self._btn_stop.clicked.connect(self._do_stop)
            self._btn_veusz.clicked.connect(self._do_open_veusz)

            for b in [self._btn_poll, self._btn_start, self._btn_stop, self._btn_veusz]:
                layout.addWidget(b)

            return widget

        def _build_status_bar(self) -> None:
            """Build the bottom status bar."""
            self._status_bar = QStatusBar()
            self.setStatusBar(self._status_bar)
            self._status_bar.showMessage(
                "Ready — configure options and click Poll Now.")

        # ----------------------------------------------------------------
        # Plot tab management
        # ----------------------------------------------------------------
        def _populate_plot_tabs(self, figures: List[Any]) -> None:
            """Clear and repopulate plot tab widget from a list of figures."""
            self._tab_widget.clear()

            if not figures:
                placeholder = QLabel(
                    "No data yet — click 'Poll Now' to fetch.")
                placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self._tab_widget.addTab(placeholder, "Waiting …")
                return

            for fig in figures:
                canvas = FigureCanvas(fig)
                toolbar = NavToolbar(canvas, self._tab_widget)
                tab_w = QWidget()
                tab_lay = QVBoxLayout(tab_w)
                tab_lay.addWidget(toolbar)
                tab_lay.addWidget(canvas)

                title = getattr(fig, "_ab_title", "Plot")
                self._tab_widget.addTab(tab_w, title[:30])

        # ----------------------------------------------------------------
        # Slots / callbacks
        # ----------------------------------------------------------------
        def _choose_out_dir(self) -> None:
            """Open a folder dialog to select the output root directory."""
            path = QFileDialog.getExistingDirectory(
                self, "Select Output Root Directory", self._le_out_dir.text()
            )
            if path:
                self._le_out_dir.setText(path)
                # Auto-update the log dir field to match the new root
                # unless the user has already customised it independently.
                default_log = os.path.join(OUTPUT_BASE_DIR, "logs")
                if self._le_log_dir.text() == default_log or not self._le_log_dir.text():
                    self._le_log_dir.setText(os.path.join(path, "logs"))

        def _choose_log_dir(self) -> None:
            """Open a folder dialog to select the log-file directory."""
            path = QFileDialog.getExistingDirectory(
                self, "Select Log Directory", self._le_log_dir.text()
            )
            if path:
                self._le_log_dir.setText(path)

        def _get_runtime_config(self) -> Dict[str, Any]:
            """Collect current GUI state into a config dict."""
            base = self._le_out_dir.text().strip() or OUTPUT_BASE_DIR
            log = self._le_log_dir.text().strip() or os.path.join(base, "logs")
            return {
                "ip_base":           IP_BASE,
                "octet_start":       self._spin_ip_start.value(),
                "octet_end":         self._spin_ip_end.value(),
                "sample_period":     self._spin_period.value(),
                # enable_gui is not surfaced in the GUI (already running);
                # it is read directly from the ENABLE_GUI module variable.
                "enable_fits":       int(self._cb_fits.isChecked()),
                "enable_csv":        int(self._cb_csv.isChecked()),
                "enable_xlsx":       int(self._cb_xlsx.isChecked()),
                "enable_log_append": int(self._cb_log.isChecked()),
                "enable_veusz":      int(self._cb_veusz.isChecked()),
                # Runtime output directories — derived from the GUI fields.
                # All sub-dirs are built under output_base_dir unless the
                # user has overridden log_dir independently.
                "output_base_dir":   base,
                "fits_dir":          os.path.join(base, "fits"),
                "csv_dir":           os.path.join(base, "csv"),
                "xlsx_dir":          os.path.join(base, "xlsx"),
                "veusz_dir":         os.path.join(base, "veusz"),
                "log_dir":           log,
            }

        def _do_poll_once(self) -> None:
            """Perform a single synchronous poll and refresh display."""
            self._append_log("Starting single poll …")
            cfg = self._get_runtime_config()
            # poll_all_devices returns a flat dict keyed by "ip_tablename".
            # update_named_dicts() converts it into the nested ALL_DEVICE_DATA
            # structure {ip: {table_name: parsed_dict}} used by all downstream
            # functions.  Always pass ALL_DEVICE_DATA to those functions, never
            # the raw flat 'data' return value.
            data = poll_all_devices(
                ip_base=cfg["ip_base"],
                octet_start=cfg["octet_start"],
                octet_end=cfg["octet_end"],
                table_names=TABLE_NAMES,
            )
            # populates ALL_DEVICE_DATA
            update_named_dicts(data)
            self._process_outputs(ALL_DEVICE_DATA, cfg)     # nested shape
            figs = build_preview_figures(ALL_DEVICE_DATA)   # nested shape
            self._populate_plot_tabs(figs)
            self._append_log(
                f"Poll complete — {len(ALL_DEVICE_DATA)} device(s), "
                f"{sum(len(t) for t in ALL_DEVICE_DATA.values())} table dicts."
            )
            self._status_bar.showMessage(
                f"Last poll: {datetime.datetime.now().strftime('%H:%M:%S')}")

        def _do_start(self) -> None:
            """Start the background polling thread."""
            if self._thread and self._thread.isRunning():
                return
            cfg = self._get_runtime_config()
            self._thread = PollThread(cfg, parent=self)
            self._thread.data_ready.connect(self._on_data_ready)
            self._thread.error_occur.connect(self._on_thread_error)
            self._thread.log_message.connect(self._append_log)
            self._thread.start()
            self._btn_start.setEnabled(False)
            self._btn_stop.setEnabled(True)
            self._status_bar.showMessage("Auto-polling started …")

        def _do_stop(self) -> None:
            """Stop the auto-poll thread, then write final Veusz file if enabled."""
            if self._thread:
                self._thread.stop()
                self._thread.wait(3000)
            self._btn_start.setEnabled(True)
            self._btn_stop.setEnabled(False)
            self._status_bar.showMessage("Auto-polling stopped.")
            # Write Veusz with all accumulated samples now that polling has stopped
            cfg = self._get_runtime_config()
            if cfg.get("enable_veusz") and ALL_DEVICE_DATA:
                veusz_dir = cfg.get("veusz_dir", VEUSZ_DIR)
                self._append_log(
                    "Writing final Veusz file with all accumulated samples…")
                try:
                    write_veusz(ALL_DEVICE_DATA, veusz_dir, show_window=False)
                    self._append_log(f"Veusz saved: {veusz_dir}")
                except Exception as exc:
                    self._append_log(f"Veusz final write error: {exc}")
                    logger.error("Veusz final write failed: %s", exc)

        def _on_data_ready(self, data: Dict) -> None:
            """Slot: called when PollThread emits new data.

            'data' here is the flat poll result; ALL_DEVICE_DATA has already
            been updated by update_named_dicts() inside PollThread.run().
            All downstream functions require the nested ALL_DEVICE_DATA shape.
            """
            cfg = self._get_runtime_config()
            self._process_outputs(ALL_DEVICE_DATA, cfg)   # nested shape
            figs = build_preview_figures(ALL_DEVICE_DATA)  # nested shape
            self._populate_plot_tabs(figs)
            self._status_bar.showMessage(
                f"Updated: {datetime.datetime.now().strftime('%H:%M:%S')}")

        def _on_thread_error(self, msg: str) -> None:
            self._append_log(f"ERROR: {msg}")
            self._status_bar.showMessage(f"Error — {msg[:60]}")

        def _do_open_veusz(self) -> None:
            """
            Open the live Veusz window (separate process window with toolbar)
            for the most recently polled data, then save as .vszh5 (HDF5).

            The output directory is read from the GUI 'Output Root Directory'
            field so it honours any runtime override made by the user.
            Runs synchronously in the GUI thread — Veusz's own event loop
            handles interaction in its separate window.  The file is saved
            when write_veusz() reaches doc.Save() after building all pages.
            """
            if not ALL_DEVICE_DATA:
                self._append_log("No data to plot — run Poll Now first.")
                return
            cfg = self._get_runtime_config()
            veusz_dir = cfg.get("veusz_dir", VEUSZ_DIR)
            self._append_log(f"Opening Veusz window … output: {veusz_dir}")
            try:
                open_veusz_preview(ALL_DEVICE_DATA, veusz_dir)
                self._append_log(f"Veusz HDF5 file(s) saved to: {veusz_dir}")
            except Exception as exc:
                self._append_log(f"Veusz error: {exc}")
                logger.error("Veusz preview failed: %s", exc)

        def _process_outputs(self, data: Dict, cfg: Dict) -> None:
            """
            Dispatch file-output writers to a background thread-pool so the
            GUI poll cycle returns immediately and sampling is never blocked
            by I/O.

            Uses flush_outputs_parallel() — the same function used by the
            headless loop.  A guard on self._flush_future prevents launching
            a new flush while the previous one is still running, which would
            risk concurrent writes to the same files.

            Veusz is intentionally excluded — it is written once on Stop
            (_do_stop) so it always contains the full accumulated dataset.

            Parameters
            ----------
            data : Dict
                Unused directly; kept for signature compatibility.  All
                writers pull from TIME_SERIES_STORE / ALL_DEVICE_DATA.
            cfg : Dict
                Runtime config dict from _get_runtime_config().
            """
            # If a previous flush is still running, skip this cycle's flush
            # rather than risk overlapping writes.  The TIME_SERIES_STORE will
            # include this cycle's data in the *next* flush that does fire.
            if self._flush_future is not None and not self._flush_future.done():
                logger.warning(
                    "Previous output flush still running — skipping flush for this cycle. "
                    "Consider increasing sample period if this recurs."
                )
                return

            self._flush_future = flush_outputs_parallel(cfg)
            # Veusz is NOT written mid-loop — it is written once when polling
            # stops (_do_stop) so it contains all accumulated samples.
            # Use 'Open in Veusz' button to preview with latest data at any time.

        def _append_log(self, text: str) -> None:
            """Append a line to the status console widget."""
            self._log_console.append(text)
            self._log_console.verticalScrollBar().setValue(
                self._log_console.verticalScrollBar().maximum()
            )

        # ----------------------------------------------------------------
        # Theme switching
        # ----------------------------------------------------------------
        def _toggle_theme(self) -> None:
            if self._dark_mode:
                self._apply_light()
            else:
                self._apply_dark()

        def _apply_dark(self) -> None:
            QtWidgets.QApplication.instance().setStyleSheet(self.DARK_STYLE)
            self._act_toggle_theme.setText("Switch to &Light Theme")
            self._dark_mode = True

        def _apply_light(self) -> None:
            QtWidgets.QApplication.instance().setStyleSheet(self.LIGHT_STYLE)
            self._act_toggle_theme.setText("Switch to &Dark Theme")
            self._dark_mode = False

        # ----------------------------------------------------------------
        # About dialog
        # ----------------------------------------------------------------
        def _show_about(self) -> None:
            from qtpy.QtWidgets import QMessageBox
            QMessageBox.information(
                self, "About AB Power Meter Monitor",
                "Allen-Bradley Site Power Meter Monitor\n"
                "NRAO / Green Bank Observatory\n\n"
                "Polls AB 1403 power meters (pages 0–10),\n"
                "stores data in named Python dicts, and\n"
                "exports to FITS, CSV, XLSX, Veusz, and logs.\n\n"
                "Author: W. Wallace\n"
                "Version: 1.0.0\n"
                "Python: 3.8+\n"
                "Qt backend: PySide6 (via QtPy)",
            )

        def closeEvent(self, event) -> None:
            """Ensure background thread and any in-flight flush stop cleanly."""
            self._do_stop()   # stops PollThread and writes final Veusz
            # Wait up to 30 s for any background file flush to complete so we
            # do not close the process while CSV / XLSX / log is still writing.
            if self._flush_future is not None and not self._flush_future.done():
                logger.info(
                    "Waiting for background flush to finish before closing…")
                try:
                    self._flush_future.result(timeout=30)
                except Exception as exc:
                    logger.error("Flush error on close: %s", exc)
            # Detach the QTextEditHandler before the widget is destroyed
            # to prevent the logging framework writing to a dangling pointer.
            if self._log_handler is not None:
                logging.getLogger("ABMonitor").removeHandler(self._log_handler)
                self._log_handler = None
            event.accept()

    # -----------------------------------------------------------------------
    # Application entry point
    # -----------------------------------------------------------------------
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication(sys.argv)

    window = MainWindow(switches=initial_switches, figures=initial_figures)
    window.show()

    sys.exit(app.exec())


# ===========================================================================
#  HEADLESS MODE — run without GUI
# ===========================================================================
# Module-level stop flag for headless loop (set by signal handler or stop script)
_HEADLESS_STOP = threading.Event()


def _install_signal_handlers() -> None:
    """
    Install SIGINT / SIGTERM handlers that set _HEADLESS_STOP cleanly.

    Also removes any pre-existing stop-signal file from the last run so a
    fresh loop does not exit immediately.
    """
    def _handler(signum, frame):  # noqa: ANN001
        logger.info(
            "Signal %s received — requesting clean stop after current poll.", signum)
        _HEADLESS_STOP.set()

    try:
        signal.signal(signal.SIGINT,  _handler)
        signal.signal(signal.SIGTERM, _handler)
    except (OSError, ValueError):
        pass   # non-main thread — signals cannot be registered; ignore

    # Remove stale stop-signal file from a previous run
    try:
        if os.path.exists(STOP_SIGNAL_FILE):
            os.remove(STOP_SIGNAL_FILE)
            logger.info("Removed stale stop-signal file: %s", STOP_SIGNAL_FILE)
    except OSError:
        pass


def run_headless(cfg: Dict[str, Any]) -> None:
    """
    Execute headless polling loop — N cycles or infinite until stopped.

    Loop behaviour
    --------------
    • HEADLESS_LOOP_COUNT = 0  → run until Ctrl-C, SIGTERM, or stop-signal
      file (``STOP_SIGNAL_FILE``) is touched.
    • HEADLESS_LOOP_COUNT = N  → run exactly N poll cycles then exit.

    Stopping cleanly
    ----------------
    From the same machine (another terminal or script)::

        touch <OUTPUT_BASE_DIR>/STOP_COLLECTION

    Or call the companion script::

        python ab_stop.py

    Or press Ctrl-C / send SIGTERM.

    The loop finishes the in-progress poll before exiting, then flushes
    all accumulated data to all enabled output formats (including Veusz).

    Adaptive flushing
    -----------------
    If TIME_SERIES_STORE grows beyond MEM_FLUSH_THRESHOLD_MB or free RAM
    falls below MEM_FREE_MIN_MB, intermediate CSV / XLSX / log writes are
    dispatched in a background thread-pool so sampling is not interrupted.
    Veusz is always written once at loop end only.

    Parameters
    ----------
    cfg : Dict[str, Any]
        Configuration dict built from module-level switch variables.
    """
    _install_signal_handlers()
    _HEADLESS_STOP.clear()   # ensure flag is clear for this run

    loop_count = cfg.get("headless_loop_count", HEADLESS_LOOP_COUNT)
    sample_sec = cfg.get("sample_period",       SAMPLE_PERIOD_SEC)
    veusz_dir = cfg.get("veusz_dir",            VEUSZ_DIR)
    infinite = (loop_count == 0)
    cycle = 0
    flush_future: Optional["concurrent.futures.Future"] = None

    logger.info(
        "Headless loop starting — %s, sample period %.1f s, output: %s",
        "infinite (stop with Ctrl-C / stop-signal file)" if infinite else f"{loop_count} cycles",
        sample_sec,
        cfg.get("output_base_dir", OUTPUT_BASE_DIR),
    )

    try:
        while True:
            # ── Check stop conditions ─────────────────────────────────────────────────────────
            if _HEADLESS_STOP.is_set():
                logger.info(
                    "Stop flag set — exiting loop after cycle %d.", cycle)
                break
            if os.path.exists(STOP_SIGNAL_FILE):
                logger.info(
                    "Stop-signal file detected (%s) — exiting.", STOP_SIGNAL_FILE)
                try:
                    os.remove(STOP_SIGNAL_FILE)
                except OSError:
                    pass
                break
            if not infinite and cycle >= loop_count:
                logger.info("Reached %d cycle(s) — loop complete.", loop_count)
                break

            # ── Poll ─────────────────────────────────────────────────────────────────
            cycle += 1
            logger.info("--- Headless cycle %d%s ---",
                        cycle,
                        f" / {loop_count}" if not infinite else "")
            t_poll_start = time.monotonic()

            data = poll_all_devices(
                ip_base=cfg["ip_base"],
                octet_start=cfg["octet_start"],
                octet_end=cfg["octet_end"],
                table_names=TABLE_NAMES,
            )
            update_named_dicts(data)  # also calls accumulate_poll()

            poll_elapsed = time.monotonic() - t_poll_start
            logger.info("Cycle %d poll complete in %.2f s — store %.1f MB, free RAM %.0f MB",
                        cycle, poll_elapsed, ts_store_size_mb(), system_free_ram_mb())

            # ── Adaptive mid-loop flush ────────────────────────────────────────────────
            # If store is getting large or RAM is tight, trigger a background
            # flush of CSV / XLSX / log (NOT Veusz — that waits for loop end).
            # Only launch a new flush if the previous one has completed.
            if should_flush(cfg):
                if flush_future is None or flush_future.done():
                    logger.info(
                        "Dispatching background flush (cycle %d)…", cycle)
                    flush_future = flush_outputs_parallel(cfg)
                else:
                    logger.warning(
                        "Previous flush still running — skipping mid-cycle flush.")

            # ── Wait for next cycle ──────────────────────────────────────────────────────────
            # Sleep in 0.5 s chunks to remain responsive to stop signals.
            remaining = sample_sec - (time.monotonic() - t_poll_start)
            while remaining > 0 and not _HEADLESS_STOP.is_set():
                if os.path.exists(STOP_SIGNAL_FILE):
                    break
                time.sleep(min(0.5, remaining))
                remaining -= 0.5

    except KeyboardInterrupt:
        logger.info("KeyboardInterrupt — stopping headless loop.")

    # ── Final flush: wait for any in-flight background flush ─────────────────────
    if flush_future is not None and not flush_future.done():
        logger.info("Waiting for background flush to complete…")
        try:
            flush_future.result(timeout=60)
        except Exception as exc:
            logger.error("Background flush error on exit: %s", exc)

    logger.info(
        "Headless loop ended after %d cycle(s). Writing final outputs…", cycle)

    # Final full write of all enabled formats (including Veusz with all samples)
    fits_dir = cfg.get("fits_dir",  FITS_DIR)
    csv_dir = cfg.get("csv_dir",   CSV_DIR)
    xlsx_dir = cfg.get("xlsx_dir",  XLSX_DIR)
    log_dir = cfg.get("log_dir",   LOG_DIR)

    if cfg.get("enable_fits"):
        write_fits(ALL_DEVICE_DATA, fits_dir)
    if cfg.get("enable_csv"):
        write_csv(ALL_DEVICE_DATA, csv_dir, append=True)
    if cfg.get("enable_xlsx"):
        write_xlsx(ALL_DEVICE_DATA, xlsx_dir)
    if cfg.get("enable_log_append"):
        write_log_text(ALL_DEVICE_DATA, log_dir)
    if cfg.get("enable_veusz"):
        logger.info(
            "Building Veusz project with all %d accumulated sample(s)…", cycle)
        write_veusz(ALL_DEVICE_DATA, veusz_dir)

    logger.info("All outputs written. Output directory: %s",
                cfg.get("output_base_dir", OUTPUT_BASE_DIR))

    # Pretty-print the 11 named dicts to stdout for inspection
    print("\n" + "="*72)
    print("  NAMED TABLE DICTS (first device — latest snapshot)")
    print("="*72)
    named = {
        "Device_Configuration_Table":           Device_Configuration_Table,
        "Communications_Configuration_Table":   Communications_Configuration_Table,
        "Voltage_Current_Table":                Voltage_Current_Table,
        "Real_Time_Power_Table":                Real_Time_Power_Table,
        "Cumulative_Power_Table":               Cumulative_Power_Table,
        "Demand_Data_Table":                    Demand_Data_Table,
        "Diagnostic_Table":                     Diagnostic_Table,
        "Voltage_Current_Snapshot_Log_Table":   Voltage_Current_Snapshot_Log_Table,
        "Power_Snapshot_Log_Table":             Power_Snapshot_Log_Table,
        "MinMax_Log_Table":                     MinMax_Log_Table,
        "Diagnostic_Table_Extended":            Diagnostic_Table_Extended,
    }
    for dname, dval in named.items():
        print(f"\n  {dname}:")
        if dval:
            print(json.dumps({k: str(v) for k, v in dval.items()}, indent=4))
        else:
            print("    (empty — no data fetched yet)")


# ===========================================================================
#  MAIN ENTRY POINT
# ===========================================================================
def main() -> None:
    """
    Application entry point.

    Reads module-level switch variables, optionally launches the GUI
    or runs a single headless poll cycle.
    """
    # Build runtime config dict from module-level switches.
    # All derived output directories are populated here from OUTPUT_BASE_DIR
    # so that run_headless() and the GUI both read from cfg rather than
    # reaching back to the module-level constants directly.
    cfg: Dict[str, Any] = {
        "ip_base":           IP_BASE,
        "octet_start":       IP_LAST_OCTET_START,
        "octet_end":         IP_LAST_OCTET_END,
        "sample_period":          SAMPLE_PERIOD_SEC,
        "headless_loop_count":    HEADLESS_LOOP_COUNT,
        "mem_flush_threshold_mb": MEM_FLUSH_THRESHOLD_MB,
        "mem_free_min_mb":        MEM_FREE_MIN_MB,
        # used by main() branch logic only; not shown in GUI
        "enable_gui":        ENABLE_GUI,
        "enable_fits":       ENABLE_FITS,
        "enable_csv":        ENABLE_CSV,
        "enable_xlsx":       ENABLE_XLSX,
        "enable_log_append": ENABLE_LOG_APPEND,
        "enable_veusz":      ENABLE_VEUSZ,   # write Veusz HDF5 project file(s)
        # Output directories — all derived from OUTPUT_BASE_DIR.
        # Change OUTPUT_BASE_DIR at the top of the file to relocate everything.
        "output_base_dir":   OUTPUT_BASE_DIR,
        "fits_dir":          FITS_DIR,
        "csv_dir":           CSV_DIR,
        "xlsx_dir":          XLSX_DIR,
        "veusz_dir":         VEUSZ_DIR,
        "log_dir":           LOG_DIR,
    }

    # Create all output directories now that cfg is finalised.
    for d in [cfg["output_base_dir"], cfg["log_dir"],
              cfg["fits_dir"], cfg["csv_dir"],
              cfg["xlsx_dir"],  cfg["veusz_dir"]]:
        os.makedirs(d, exist_ok=True)

    logger.info("AB Power Meter Monitor starting up.")
    logger.info("Configuration: %s", json.dumps(cfg, indent=2))

    if ENABLE_GUI:
        # Pre-compute an initial (empty) figure set; GUI will refresh on first poll
        initial_figs: List[Any] = []
        try:
            launch_gui(initial_switches=cfg, initial_figures=initial_figs)
        except Exception as exc:
            logger.critical("GUI launch failed: %s\n%s",
                            exc, traceback.format_exc())
            logger.info("Falling back to headless mode.")
            run_headless(cfg)
    else:
        run_headless(cfg)


if __name__ == "__main__":
    main()
